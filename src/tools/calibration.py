#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""关键词实验数据采集工具。"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import openpyxl

from src.core import should_stop, wait_if_paused
from src.version import __version__

try:
    from src.platforms.tiktok.keyword import run_tiktok_spider
except ModuleNotFoundError:
    run_tiktok_spider = None

try:
    from src.platforms.x_twitter.keyword import run_x_spider
except ModuleNotFoundError:
    run_x_spider = None

try:
    from src.platforms.youtube.keyword import run_youtube_spider
except ModuleNotFoundError:
    run_youtube_spider = None


VALID_PLATFORMS = ("youtube", "tiktok", "x_twitter")
DEFAULT_TRACK_LANGUAGE = "default"
KEYWORD_ROLE_OFFICIAL = "official"
KEYWORD_ROLE_CANDIDATE = "candidate"

STATUS_SUCCESS = "SUCCESS"
STATUS_EMPTY_RESULT = "EMPTY_RESULT"
STATUS_AUTH_REQUIRED = "AUTH_REQUIRED"
STATUS_CAPTCHA_OR_RISK = "CAPTCHA_OR_RISK"
STATUS_QUOTA_EXCEEDED = "QUOTA_EXCEEDED"
STATUS_TIMEOUT = "TIMEOUT"
STATUS_OUTPUT_SCHEMA_ERROR = "OUTPUT_SCHEMA_ERROR"
STATUS_UNKNOWN_PLATFORM = "UNKNOWN_PLATFORM"
STATUS_FAILED = "FAILED"

STANDARDIZED_SHEET_NAME = "标准化采集数据"
STANDARDIZED_HEADERS = [
    "游戏名",
    "平台",
    "语言",
    "关键词角色",
    "关键词文本",
    "内容ID",
    "内容链接",
    "标题",
    "简介",
    "发布时间",
    "原始输出文件",
    "采集状态",
]

SHEET_NAME_ALIASES = {
    "youtube": ("视频信息", "数据"),
    "tiktok": ("视频信息", "数据"),
    "x_twitter": ("数据", "推文信息"),
}

PLATFORM_HEADER_ALIASES = {
    "youtube": {
        "link": ("视频链接", "链接", "url", "link"),
        "title": ("视频标题", "标题"),
        "description": ("视频简介", "视频描述", "简介", "描述"),
        "published_at": ("发布时间", "发布日期", "发布时", "发布时间戳"),
    },
    "tiktok": {
        "link": ("视频链接", "链接", "url", "link"),
        "title": ("视频标题", "标题", "内容"),
        "description": ("视频简介", "视频描述", "简介", "描述"),
        "published_at": ("发布时间", "发布日期", "发布时"),
    },
    "x_twitter": {
        "link": ("推文链接", "链接", "url", "link"),
        "title": ("标题",),
        "description": ("推文内容", "内容", "正文", "简介", "描述"),
        "published_at": ("发帖时间", "发布时间", "发布日期"),
    },
}

_NON_WORD_RE = re.compile(r"[\W_]+", re.UNICODE)


@dataclass(frozen=True)
class StandardizedRecord:
    game_name: str
    platform: str
    language: str
    keyword_role: str
    keyword_text: str
    content_id: str
    content_link: str
    title: str
    description: str
    published_at: str
    source_output_file: str
    collection_status: str

    def to_row(self) -> list[str]:
        return [
            self.game_name,
            self.platform,
            self.language,
            self.keyword_role,
            self.keyword_text,
            self.content_id,
            self.content_link,
            self.title,
            self.description,
            self.published_at,
            self.source_output_file,
            self.collection_status,
        ]


@dataclass
class SpiderRunResult:
    platform: str
    keyword: str
    status: str
    ids: set[str]
    links: set[str]
    output_path: str | None
    error_message: str | None
    started_at: str
    finished_at: str
    scanned_count: int | None = None
    written_count: int | None = None
    hit_limit: bool = False
    records: list[dict[str, str]] = field(default_factory=list)

    def to_snapshot(self) -> dict[str, Any]:
        return {
            "platform": self.platform,
            "keyword": self.keyword,
            "status": self.status,
            "ids": sorted(self.ids),
            "links": sorted(self.links),
            "output_path": self.output_path,
            "error_message": self.error_message,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "scanned_count": self.scanned_count,
            "written_count": self.written_count,
            "hit_limit": self.hit_limit,
            "record_count": len(self.records),
        }


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def workspace_root() -> Path:
    return Path(__file__).resolve().parents[2]


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r'[\\/*?:"<>|]', "", str(value or "")).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:80] or "item"


def raw_game_dir_name(game_name: str, game_index: int) -> str:
    return f"{game_index:02d}_{sanitize_filename_part(game_name)}"


def raw_platform_dir_name(platform: str) -> str:
    return sanitize_filename_part(platform)


def raw_language_dir_name(language: str) -> str:
    return sanitize_filename_part(language or DEFAULT_TRACK_LANGUAGE)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_platforms(platforms_cfg: Any) -> list[str]:
    if platforms_cfg is None:
        return list(VALID_PLATFORMS)

    if isinstance(platforms_cfg, list):
        raw_platforms = [item for item in platforms_cfg if isinstance(item, str)]
    elif isinstance(platforms_cfg, str):
        raw_platforms = platforms_cfg.split(",")
    else:
        raw_platforms = []

    normalized: list[str] = []
    seen: set[str] = set()
    for raw in raw_platforms:
        name = raw.strip().lower()
        if not name or name in seen:
            continue
        seen.add(name)
        normalized.append(name)
    return normalized or list(VALID_PLATFORMS)


def invalid_platforms(platforms: list[str]) -> list[str]:
    return [platform for platform in platforms if platform not in VALID_PLATFORMS]


def parse_keyword_groups_text(raw_text: str) -> list[list[str]]:
    groups: list[list[str]] = []
    for raw_line in (raw_text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        keywords = [item.strip() for item in re.split(r"[,，]", line) if item.strip()]
        if keywords:
            groups.append(keywords)
    return groups


def format_keyword_groups_text(keyword_groups: list[list[str]]) -> str:
    lines: list[str] = []
    for group in keyword_groups or []:
        keywords = [str(keyword).strip() for keyword in group if str(keyword).strip()]
        if keywords:
            lines.append(", ".join(keywords))
    return "\n".join(lines)


def parse_keyword_list_text(raw_text: str) -> list[str]:
    keywords: list[str] = []
    seen: set[str] = set()
    for raw_line in (raw_text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        for item in re.split(r"[,，]", line):
            keyword = item.strip()
            lowered = keyword.casefold()
            if not keyword or lowered in seen:
                continue
            seen.add(lowered)
            keywords.append(keyword)
    return keywords


def format_keyword_list_text(keywords: list[str]) -> str:
    return "\n".join(str(keyword).strip() for keyword in keywords or [] if str(keyword).strip())


def normalize_track_language(language: Any) -> str:
    text = str(language or "").strip().lower()
    return text or DEFAULT_TRACK_LANGUAGE


def build_track_key(platform: str, language: str) -> str:
    return f"{platform}/{language}"


def normalize_keyword_list(
    raw_keywords: Any,
    *,
    game_index: int,
    track_index: int,
    field_name: str,
    allow_empty: bool,
) -> list[str]:
    if raw_keywords is None:
        if allow_empty:
            return []
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 缺少 {field_name}。")
    if not isinstance(raw_keywords, list):
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 的 {field_name} 必须是数组。")

    normalized: list[str] = []
    seen: set[str] = set()
    for raw_keyword in raw_keywords:
        keyword = str(raw_keyword or "").strip()
        lowered = keyword.casefold()
        if not keyword or lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(keyword)

    if not normalized and not allow_empty:
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 的 {field_name} 不能为空。")
    return normalized


def flatten_legacy_keyword_groups(
    raw_groups: Any,
    *,
    game_index: int,
    track_index: int,
) -> list[str]:
    if raw_groups is None:
        return []
    if not isinstance(raw_groups, list):
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 的 keyword_groups 必须是数组。")

    flattened: list[str] = []
    seen: set[str] = set()
    for group_index, group in enumerate(raw_groups, 1):
        if not isinstance(group, list):
            raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 的第 {group_index} 个词组必须是数组。")
        for raw_keyword in group:
            keyword = str(raw_keyword or "").strip()
            lowered = keyword.casefold()
            if not keyword or lowered in seen:
                continue
            seen.add(lowered)
            flattened.append(keyword)
    return flattened


def normalize_track_config(track: dict[str, Any], *, game_index: int, track_index: int) -> dict[str, Any]:
    if not isinstance(track, dict):
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 必须是对象。")

    platform = str(track.get("platform", "")).strip().lower()
    language = normalize_track_language(track.get("language", DEFAULT_TRACK_LANGUAGE))

    if not platform:
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 缺少 platform。")
    if platform not in VALID_PLATFORMS:
        raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 使用了不支持的平台: {platform}")

    if "official_keywords" in track or "candidate_keywords" in track:
        official_keywords = normalize_keyword_list(
            track.get("official_keywords"),
            game_index=game_index,
            track_index=track_index,
            field_name="official_keywords",
            allow_empty=False,
        )
        candidate_keywords = normalize_keyword_list(
            track.get("candidate_keywords", []),
            game_index=game_index,
            track_index=track_index,
            field_name="candidate_keywords",
            allow_empty=True,
        )
    else:
        baseline_query = str(track.get("baseline_query", "")).strip()
        if not baseline_query:
            raise ValueError(f"第 {game_index} 个游戏的第 {track_index} 个 track 缺少 baseline_query。")
        official_keywords = [baseline_query]
        candidate_keywords = flatten_legacy_keyword_groups(
            track.get("keyword_groups", []),
            game_index=game_index,
            track_index=track_index,
        )

    return {
        "platform": platform,
        "language": language,
        "official_keywords": official_keywords,
        "candidate_keywords": candidate_keywords,
    }


def expand_legacy_game_tracks(
    *,
    baseline_query: str,
    raw_groups: Any,
    game_index: int,
) -> list[dict[str, Any]]:
    candidate_keywords = flatten_legacy_keyword_groups(raw_groups, game_index=game_index, track_index=1)
    return [
        {
            "platform": platform,
            "language": DEFAULT_TRACK_LANGUAGE,
            "official_keywords": [baseline_query],
            "candidate_keywords": list(candidate_keywords),
        }
        for platform in VALID_PLATFORMS
    ]


def normalize_games_config(games: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_games: list[dict[str, Any]] = []
    for game_index, game in enumerate(games, 1):
        if not isinstance(game, dict):
            raise ValueError(f"第 {game_index} 个游戏配置必须是对象。")

        name = str(game.get("name", "")).strip()
        if not name:
            raise ValueError(f"第 {game_index} 个游戏缺少名称。")

        raw_tracks = game.get("tracks")
        if raw_tracks is None:
            baseline_query = str(game.get("baseline_query", "")).strip()
            if not baseline_query:
                raise ValueError(f"第 {game_index} 个游戏缺少 baseline_query。")
            tracks = expand_legacy_game_tracks(
                baseline_query=baseline_query,
                raw_groups=game.get("keyword_groups", []),
                game_index=game_index,
            )
        else:
            if not isinstance(raw_tracks, list) or not raw_tracks:
                raise ValueError(f"第 {game_index} 个游戏至少需要一个 track。")
            tracks = [
                normalize_track_config(track, game_index=game_index, track_index=track_index)
                for track_index, track in enumerate(raw_tracks, 1)
            ]

        seen_track_keys: set[str] = set()
        for track in tracks:
            track_key = build_track_key(track["platform"], track["language"])
            if track_key in seen_track_keys:
                raise ValueError(f"第 {game_index} 个游戏存在重复 track: {track_key}")
            seen_track_keys.add(track_key)

        normalized_games.append({"name": name, "tracks": tracks})

    if not normalized_games:
        raise ValueError("请至少配置一个游戏。")
    return normalized_games


def parse_games_definition(raw_definition: str) -> list[dict[str, Any]]:
    text = (raw_definition or "").strip()
    if not text:
        raise ValueError("请至少配置一个游戏。")

    if text.startswith("[") or text.startswith("{"):
        try:
            payload = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"JSON 游戏配置解析失败: {exc}") from exc
        if isinstance(payload, dict):
            payload = payload.get("games")
        if not isinstance(payload, list):
            raise ValueError("JSON 游戏配置必须是数组，或包含 games 数组的对象。")
        return normalize_games_config(payload)

    blocks = re.split(r"\n\s*\n+", text)
    games: list[dict[str, Any]] = []
    for block_index, block in enumerate(blocks, 1):
        lines: list[str] = []
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            lines.append(line)
        if not lines:
            continue

        header = lines[0]
        if "|" not in header:
            raise ValueError(f"第 {block_index} 个游戏配置的首行必须写成“游戏名称 | 官方关键词”。")
        name, official_keyword = [part.strip() for part in header.split("|", 1)]
        if not name or not official_keyword:
            raise ValueError(f"第 {block_index} 个游戏配置的名称和官方关键词都不能为空。")
        games.append(
            {
                "name": name,
                "baseline_query": official_keyword,
                "keyword_groups": parse_keyword_groups_text("\n".join(lines[1:])),
            }
        )
    return normalize_games_config(games)


def select_matching_tracks(games: list[dict[str, Any]], platforms: list[str]) -> list[tuple[str, dict[str, Any]]]:
    matched_tracks: list[tuple[str, dict[str, Any]]] = []
    for game in games:
        game_name = str(game.get("name", "")).strip()
        for track in game.get("tracks", []):
            if track.get("platform") in platforms:
                matched_tracks.append((game_name, track))
    return matched_tracks


def validate_selected_platforms(games: list[dict[str, Any]], platforms: list[str]) -> None:
    invalid = invalid_platforms(platforms)
    if invalid:
        raise ValueError(f"不支持的平台: {', '.join(invalid)}")
    if not select_matching_tracks(games, platforms):
        raise ValueError("所选平台与实验配置中的 track 不匹配，请检查平台选择或 track 配置。")


def extract_id_from_link(link: str, platform: str) -> str:
    if not link or not isinstance(link, str):
        return ""

    text = link.strip()
    try:
        if platform == "youtube":
            match = re.search(r"(?:v=|/shorts/|/embed/|/live/|/v/|youtu\.be/)([a-zA-Z0-9_-]{11})", text)
            if match:
                return match.group(1)
            parsed = urlparse(text)
            if parsed.query:
                query = parse_qs(parsed.query)
                if query.get("v"):
                    return query["v"][0]
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]

        if platform == "tiktok":
            match = re.search(r"/video/(\d+)", text)
            if match:
                return match.group(1)
            match = re.search(r"/v/(\d+)(?:\.html)?", text)
            if match:
                return match.group(1)
            parsed = urlparse(text)
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]

        if platform == "x_twitter":
            match = re.search(r"/status/(\d+)", text)
            if match:
                return match.group(1)
            parsed = urlparse(text)
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]
    except Exception:
        return ""

    return text


def load_config(config_path: str) -> dict[str, Any]:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    with open(config_path, "r", encoding="utf-8") as file:
        config = json.load(file)

    if "games" not in config or not isinstance(config["games"], list):
        raise ValueError("Configuration must contain a 'games' list.")

    config["games"] = normalize_games_config(config["games"])
    return config


def select_workbook_sheet(workbook: openpyxl.Workbook, platform: str):
    for name in SHEET_NAME_ALIASES.get(platform, ()):
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook.active


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return _NON_WORD_RE.sub("", stringify_cell(value).casefold())


def find_column_index(headers: list[Any], aliases: tuple[str, ...]) -> int | None:
    normalized_headers = [normalize_header(header) for header in headers]
    normalized_aliases = [normalize_header(alias) for alias in aliases if normalize_header(alias)]

    for alias in normalized_aliases:
        if alias in normalized_headers:
            return normalized_headers.index(alias)

    for index, header in enumerate(normalized_headers):
        if any(alias and alias in header for alias in normalized_aliases):
            return index
    return None


def extract_records_from_excel(file_path: str, platform: str) -> list[dict[str, str]]:
    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = select_workbook_sheet(workbook, platform)
        try:
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        except StopIteration:
            return []

        platform_aliases = PLATFORM_HEADER_ALIASES.get(platform, {})
        link_index = find_column_index(headers, platform_aliases.get("link", ()))
        if link_index is None:
            raise ValueError("Link column not found.")

        title_index = find_column_index(headers, platform_aliases.get("title", ()))
        description_index = find_column_index(headers, platform_aliases.get("description", ()))
        published_at_index = find_column_index(headers, platform_aliases.get("published_at", ()))

        records: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= link_index:
                continue
            content_link = stringify_cell(row[link_index])
            if not content_link:
                continue

            title = stringify_cell(row[title_index]) if title_index is not None and len(row) > title_index else ""
            description = (
                stringify_cell(row[description_index]) if description_index is not None and len(row) > description_index else ""
            )
            published_at = (
                stringify_cell(row[published_at_index]) if published_at_index is not None and len(row) > published_at_index else ""
            )

            if platform == "x_twitter" and description and not title:
                title = ""
            if platform in {"youtube", "tiktok"} and not description:
                description = ""

            records.append(
                {
                    "content_id": extract_id_from_link(content_link, platform),
                    "content_link": content_link,
                    "title": title,
                    "description": description,
                    "published_at": published_at,
                }
            )
        return records
    finally:
        if workbook is not None:
            workbook.close()


def classify_error_status(message: str | None) -> str:
    lowered = (message or "").lower()
    if any(token in lowered for token in ("quota", "rate limit", "too many requests", "配额")):
        return STATUS_QUOTA_EXCEEDED
    if any(token in lowered for token in ("captcha", "verify", "risk", "风控")):
        return STATUS_CAPTCHA_OR_RISK
    if any(token in lowered for token in ("login", "sign in", "signin", "auth", "unauthorized", "forbidden", "permission", "登录")):
        return STATUS_AUTH_REQUIRED
    if any(token in lowered for token in ("timeout", "timed out", "超时")):
        return STATUS_TIMEOUT
    return STATUS_FAILED


def select_x_search_tab(platform_config: dict[str, Any]) -> str:
    search_tab = str(platform_config.get("x_search_tab", "latest") or "latest").strip().lower()
    if search_tab not in {"latest", "top"}:
        return "latest"
    return search_tab


def run_platform_spider(
    platform: str,
    keyword: str,
    start_date: str,
    end_date: str,
    platform_config: dict[str, Any],
    days: int,
    stop_event=None,
    pause_event=None,
) -> SpiderRunResult:
    retrieved_path: str | None = None
    spider_stats: dict[str, Any] = {}
    started_at = now_str()

    def finish_callback(path):
        nonlocal retrieved_path
        retrieved_path = path

    def stats_callback(payload: dict[str, Any]):
        nonlocal spider_stats
        spider_stats = dict(payload or {})

    def log_callback(message: str):
        logging.debug("[keyword_collection] %s", message)

    try:
        if platform == "youtube":
            if run_youtube_spider is None:
                raise ModuleNotFoundError("google-api-python-client is required for YouTube keyword scraping")
            run_youtube_spider(
                api_keys=platform_config.get("api_keys", []),
                keywords_list=[keyword],
                max_results=int(platform_config.get("max_results", 10)),
                limit_time_str="是",
                start_date=start_date,
                end_date=end_date,
                get_comments_str="否",
                max_comments=0,
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
                config={"youtube_search_method": "使用 API（消耗配额）"},
                stats_callback=stats_callback,
            )
        elif platform == "tiktok":
            if run_tiktok_spider is None:
                raise ModuleNotFoundError("playwright is required for TikTok keyword scraping")
            max_videos = int(platform_config.get("max_videos", 10))
            run_tiktok_spider(
                keywords_list=[keyword],
                max_videos=max_videos,
                max_candidates=max(int(platform_config.get("max_candidates", max_videos * 3)), max_videos),
                limit_time_str="是",
                start_date=start_date,
                end_date=end_date,
                get_comments_str="否",
                max_comments=0,
                cdp_port_or_url=platform_config.get("cdp_url", "http://localhost:9222"),
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
                stats_callback=stats_callback,
            )
        elif platform == "x_twitter":
            if run_x_spider is None:
                raise ModuleNotFoundError("playwright is required for X keyword scraping")
            adv_params = {
                "limit_time": "是",
                "start_date": start_date,
                "end_date": end_date,
                "get_comments": "否",
                "max_comments": 0,
                "lang": "any",
                "search_tab": "live" if select_x_search_tab(platform_config) == "latest" else "top",
            }
            run_x_spider(
                keywords_list=[keyword],
                adv_params=adv_params,
                port=platform_config.get("cdp_url", "http://localhost:9222"),
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
                config={
                    "max_scrolls": int(platform_config.get("max_scrolls", 2)),
                    "cooldown_min": 2.0,
                    "cooldown_max": 4.0,
                    "no_new_scroll_limit": 2,
                    "slice_days": days,
                    "max_parallel_tabs": 1,
                },
                stats_callback=stats_callback,
            )
        else:
            finished_at = now_str()
            return SpiderRunResult(
                platform=platform,
                keyword=keyword,
                status=STATUS_UNKNOWN_PLATFORM,
                ids=set(),
                links=set(),
                output_path=None,
                error_message=f"Unknown platform: {platform}",
                started_at=started_at,
                finished_at=finished_at,
            )
    except Exception as exc:
        finished_at = now_str()
        error_message = str(exc)
        logging.exception("爬虫执行异常 (platform=%s, keyword=%s)", platform, keyword)
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=classify_error_status(error_message),
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=error_message,
            started_at=started_at,
            finished_at=finished_at,
            scanned_count=spider_stats.get("scanned_count"),
            written_count=spider_stats.get("written_count"),
            hit_limit=bool(spider_stats.get("hit_limit", False)),
        )

    finished_at = now_str()
    if not retrieved_path:
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_FAILED,
            ids=set(),
            links=set(),
            output_path=None,
            error_message="No Excel path returned from spider",
            started_at=started_at,
            finished_at=finished_at,
            scanned_count=spider_stats.get("scanned_count"),
            written_count=spider_stats.get("written_count"),
            hit_limit=bool(spider_stats.get("hit_limit", False)),
        )

    if not os.path.exists(retrieved_path):
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_OUTPUT_SCHEMA_ERROR,
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=f"Excel file not found at {retrieved_path}",
            started_at=started_at,
            finished_at=finished_at,
            scanned_count=spider_stats.get("scanned_count"),
            written_count=spider_stats.get("written_count"),
            hit_limit=bool(spider_stats.get("hit_limit", False)),
        )

    try:
        records = extract_records_from_excel(retrieved_path, platform)
    except Exception as exc:
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_OUTPUT_SCHEMA_ERROR,
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=f"Failed to parse Excel: {exc}",
            started_at=started_at,
            finished_at=finished_at,
            scanned_count=spider_stats.get("scanned_count"),
            written_count=spider_stats.get("written_count"),
            hit_limit=bool(spider_stats.get("hit_limit", False)),
        )

    links = {record["content_link"] for record in records if record.get("content_link")}
    ids = {record["content_id"] for record in records if record.get("content_id")}
    status = STATUS_SUCCESS if ids else STATUS_EMPTY_RESULT
    return SpiderRunResult(
        platform=platform,
        keyword=keyword,
        status=status,
        ids=ids,
        links=links,
        output_path=retrieved_path,
        error_message="",
        started_at=started_at,
        finished_at=finished_at,
        scanned_count=spider_stats.get("scanned_count"),
        written_count=spider_stats.get("written_count", len(records)),
        hit_limit=bool(spider_stats.get("hit_limit", False)),
        records=records,
    )


def resolve_output_base(output_path: str, log_callback=None) -> Path:
    if not output_path:
        return workspace_root() / "output" / "calibration"

    raw_path = Path(output_path)
    if raw_path.suffix:
        if log_callback:
            log_callback(f"检测到旧版文件路径输出，已按目录模式处理: {raw_path}")
        base_path = raw_path.parent / "calibration"
    else:
        base_path = raw_path

    if not base_path.is_absolute():
        base_path = workspace_root() / base_path
    return base_path


def create_run_directory(output_path: str, log_callback=None) -> tuple[str, Path]:
    base_path = resolve_output_base(output_path, log_callback=log_callback)
    run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    run_dir = base_path / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_id, run_dir


def build_environment_snapshot(
    *,
    run_id: str,
    run_started_at: str,
    run_finished_at: str,
    start_date: str,
    end_date: str,
    days: int,
    platforms: list[str],
    x_search_tab: str,
    standardized_output_path: str,
) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "start_date": start_date,
        "end_date": end_date,
        "days": days,
        "platforms": platforms,
        "x_search_tab": x_search_tab,
        "tool_version": __version__,
        "started_at": run_started_at,
        "finished_at": run_finished_at,
        "standardized_output_path": standardized_output_path,
    }


def build_keyword_snapshot(
    *,
    game_name: str,
    platform: str,
    language: str,
    track_key: str,
    keyword_role: str,
    keyword: str,
    result: SpiderRunResult,
) -> dict[str, Any]:
    payload = result.to_snapshot()
    payload.update(
        {
            "game_name": game_name,
            "platform": platform,
            "language": language,
            "track_key": track_key,
            "keyword_role": keyword_role,
            "keyword": keyword,
        }
    )
    return payload


def build_standardized_rows(
    *,
    game_name: str,
    platform: str,
    language: str,
    keyword_role: str,
    keyword_text: str,
    result: SpiderRunResult,
) -> list[StandardizedRecord]:
    if result.records:
        return [
            StandardizedRecord(
                game_name=game_name,
                platform=platform,
                language=language,
                keyword_role=keyword_role,
                keyword_text=keyword_text,
                content_id=record.get("content_id", ""),
                content_link=record.get("content_link", ""),
                title=record.get("title", ""),
                description=record.get("description", ""),
                published_at=record.get("published_at", ""),
                source_output_file=result.output_path or "",
                collection_status=result.status,
            )
            for record in result.records
        ]

    return [
        StandardizedRecord(
            game_name=game_name,
            platform=platform,
            language=language,
            keyword_role=keyword_role,
            keyword_text=keyword_text,
            content_id="",
            content_link="",
            title="",
            description="",
            published_at="",
            source_output_file=result.output_path or "",
            collection_status=result.status,
        )
    ]


def write_standardized_workbook(rows: list[StandardizedRecord], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = STANDARDIZED_SHEET_NAME
    worksheet.append(STANDARDIZED_HEADERS)
    for record in rows:
        worksheet.append(record.to_row())
    worksheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def run_calibration_task(config: dict[str, Any], output_path: str, log_callback=None, stop_event=None, pause_event=None) -> str:
    time_period = config.get("time_period", {})
    days_raw = time_period.get("days", 7)
    try:
        days = int(days_raw)
    except (TypeError, ValueError) as exc:
        if log_callback:
            log_callback(f"Invalid days value in config (must be integer): {exc}")
        raise ValueError(f"Invalid days value: {exc}") from exc

    if "start_date" in time_period and "end_date" in time_period:
        start_date_str = str(time_period["start_date"])
        end_date_str = str(time_period["end_date"])
        try:
            start_dt = dt.datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = dt.datetime.strptime(end_date_str, "%Y-%m-%d")
            days = max(1, (end_dt - start_dt).days)
        except (TypeError, ValueError):
            pass
    else:
        end_dt = dt.datetime.now()
        start_dt = end_dt - dt.timedelta(days=days)
        start_date_str = start_dt.strftime("%Y-%m-%d")
        end_date_str = end_dt.strftime("%Y-%m-%d")

    platforms = parse_platforms(config.get("platforms"))
    normalized_games = normalize_games_config(config.get("games", []))
    validate_selected_platforms(normalized_games, platforms)

    run_started_at = now_str()
    run_id, run_dir = create_run_directory(output_path, log_callback=log_callback)
    x_search_tab = select_x_search_tab(config.get("x_twitter", {}))

    if log_callback:
        log_callback(f"采集时间范围: {start_date_str} 到 {end_date_str} ({days} 天)")
        log_callback(f"运行目录: {run_dir}")

    snapshot_config = dict(config)
    snapshot_config["games"] = normalized_games
    write_json(run_dir / "config_snapshot.json", snapshot_config)

    collection_rows: list[StandardizedRecord] = []

    for game_index, game in enumerate(normalized_games, 1):
        if should_stop(stop_event):
            break
        if wait_if_paused(pause_event, stop_event):
            break

        game_name = game["name"]
        matched_tracks = [track for track in game["tracks"] if track["platform"] in platforms]
        if log_callback:
            log_callback(f"\n处理游戏: {game_name}")
            if not matched_tracks:
                log_callback("  所选平台下没有匹配 track，跳过。")

        for track in matched_tracks:
            if should_stop(stop_event):
                break
            if wait_if_paused(pause_event, stop_event):
                break

            platform = track["platform"]
            language = track["language"]
            track_key = build_track_key(platform, language)
            platform_config = dict(config.get(platform, {}))
            if platform == "x_twitter":
                platform_config["x_search_tab"] = x_search_tab

            raw_dir = run_dir / "raw" / raw_game_dir_name(game_name, game_index) / raw_platform_dir_name(platform) / raw_language_dir_name(language)

            if log_callback:
                log_callback(f"  运行 track: {track_key}")

            for keyword_role, keywords in (
                (KEYWORD_ROLE_OFFICIAL, list(track.get("official_keywords", []))),
                (KEYWORD_ROLE_CANDIDATE, list(track.get("candidate_keywords", []))),
            ):
                for keyword_index, keyword in enumerate(keywords, 1):
                    if should_stop(stop_event):
                        break
                    if wait_if_paused(pause_event, stop_event):
                        break

                    if log_callback:
                        log_callback(f"    [{keyword_role}] {keyword}")

                    result = run_platform_spider(
                        platform=platform,
                        keyword=keyword,
                        start_date=start_date_str,
                        end_date=end_date_str,
                        platform_config=platform_config,
                        days=days,
                        stop_event=stop_event,
                        pause_event=pause_event,
                    )

                    snapshot_name = f"{keyword_role}_{keyword_index:02d}.json"
                    write_json(
                        raw_dir / snapshot_name,
                        build_keyword_snapshot(
                            game_name=game_name,
                            platform=platform,
                            language=language,
                            track_key=track_key,
                            keyword_role=keyword_role,
                            keyword=keyword,
                            result=result,
                        ),
                    )
                    collection_rows.extend(
                        build_standardized_rows(
                            game_name=game_name,
                            platform=platform,
                            language=language,
                            keyword_role=keyword_role,
                            keyword_text=keyword,
                            result=result,
                        )
                    )

    reports_dir = run_dir / "reports"
    standardized_output_path = reports_dir / "keyword_collection_standardized.xlsx"
    write_standardized_workbook(collection_rows, standardized_output_path)

    run_finished_at = now_str()
    write_json(
        run_dir / "environment_snapshot.json",
        build_environment_snapshot(
            run_id=run_id,
            run_started_at=run_started_at,
            run_finished_at=run_finished_at,
            start_date=start_date_str,
            end_date=end_date_str,
            days=days,
            platforms=platforms,
            x_search_tab=x_search_tab,
            standardized_output_path=str(standardized_output_path),
        ),
    )

    if log_callback:
        log_callback(f"标准化采集表: {standardized_output_path}")
        log_callback(f"运行完成: {run_dir}")

    return str(run_dir)


def main() -> None:
    parser = argparse.ArgumentParser(description="Keyword collection tool")
    parser.add_argument("--config", type=str, default="config/calibration_config.json", help="Path to configuration file")
    parser.add_argument("--output", type=str, default="output/calibration", help="Output root or legacy report file path")
    args = parser.parse_args()

    try:
        config = load_config(args.config)
    except Exception as exc:
        print(f"Failed to load config: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        run_calibration_task(config, args.output)
    except Exception as exc:
        print(f"Keyword collection failed: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
