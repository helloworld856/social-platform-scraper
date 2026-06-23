from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import openpyxl
from langchain_core.messages import HumanMessage, SystemMessage

from src.core import build_chat_openai, interruptible_sleep, should_stop, wait_if_paused

OUTPUT_HEADERS = [
    "AI判定词",
    "AI判定列",
    "首轮判定",
    "首轮理由",
    "复判判定",
    "复判理由",
    "最终判定",
    "复核状态",
    "异常信息",
]

PASS_RESULT_MATCH = {"符合", "不符合"}
PASS_RESULT_EMPTY = "空内容"
PASS_RESULT_FAILED = "失败"
FINAL_RESULT_PENDING = "待人工复核"
REVIEW_STATUS_CONSISTENT = "一致"
REVIEW_STATUS_PENDING = "待人工复核"
REVIEW_STATUS_EMPTY = "空内容"

SYSTEM_PROMPT = """
你是一个内容语义规则判定助手。

输入是一组行号和文本内容，以及一个“判定词/判定规则”。
你需要根据文本内容本身，判断每一行是否“符合”该规则。

输出必须是标准 JSON 数组，不要输出 Markdown，不要输出解释，不要输出代码块。
数组中的每个元素必须是一个对象，格式如下：
[
  {"行号": 2, "判定结果": "符合", "简短理由": "一句简短理由"},
  {"行号": 3, "判定结果": "不符合", "简短理由": "一句简短理由"}
]

规则：
1. “判定结果”只能是“符合”或“不符合”。
2. “简短理由”必须简洁，控制在一小句内。
3. 必须严格保持输入行号顺序，不要新增、删除、合并任何记录。
4. 只根据提供的文本内容判断，不要臆测未提供的信息。
"""


@dataclass(frozen=True)
class RowPayload:
    row_number: int
    content: str


@dataclass
class RowState:
    rule_text: str
    target_columns_text: str
    pass1_result: str = ""
    pass1_reason: str = ""
    pass2_result: str = ""
    pass2_reason: str = ""
    final_result: str = ""
    review_status: str = ""
    error_message: str = ""


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def extract_sheet_names(input_xlsx: str | Path) -> list[str]:
    workbook = openpyxl.load_workbook(input_xlsx, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def extract_sheet_headers(input_xlsx: str | Path, sheet_name: str) -> list[str]:
    workbook = openpyxl.load_workbook(input_xlsx, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet 不存在：{sheet_name}")
        worksheet = workbook[sheet_name]
        return validate_headers([cell.value for cell in worksheet[1]])
    finally:
        workbook.close()


def validate_headers(header_values: list[Any]) -> list[str]:
    headers = [normalize_header(value) for value in header_values]
    if not headers or not any(headers):
        raise ValueError("目标 sheet 的表头为空，无法选择判定列。")

    duplicates: set[str] = set()
    seen: set[str] = set()
    for header in headers:
        if not header:
            raise ValueError("目标 sheet 存在空表头，请先补齐表头。")
        lowered = header.casefold()
        if lowered in seen:
            duplicates.add(header)
        seen.add(lowered)
    if duplicates:
        duplicate_text = "、".join(sorted(duplicates))
        raise ValueError(f"目标 sheet 存在重复表头：{duplicate_text}")
    return headers


def resolve_target_columns(headers: list[str], selected_columns: list[str]) -> list[str]:
    selected = {normalize_header(value).casefold() for value in selected_columns if normalize_header(value)}
    ordered = [header for header in headers if header.casefold() in selected]
    if not ordered:
        raise ValueError("至少需要选择 1 个判定列。")
    if len(ordered) != len(selected):
        missing = [value for value in selected_columns if normalize_header(value).casefold() not in {header.casefold() for header in ordered}]
        raise ValueError(f"存在无效的判定列：{', '.join(missing)}")
    return ordered


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def build_row_content(headers: list[str], row_values: list[Any], target_columns: list[str]) -> str:
    target_set = {header.casefold() for header in target_columns}
    parts: list[str] = []
    for header, value in zip(headers, row_values):
        if header.casefold() not in target_set:
            continue
        text = stringify_cell(value)
        if not text:
            continue
        parts.append(f"{header}：{text}")
    return "\n\n".join(parts)


def clean_json_text(text: str) -> str:
    cleaned = str(text or "").strip()
    if cleaned.startswith("```json"):
        cleaned = cleaned[7:].strip()
    elif cleaned.startswith("```"):
        cleaned = cleaned[3:].strip()
    if cleaned.endswith("```"):
        cleaned = cleaned[:-3].strip()
    return cleaned


def extract_message_text(response: Any) -> str:
    content = getattr(response, "content", response)
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
                continue
            if isinstance(item, dict):
                parts.append(str(item.get("text", "")))
                continue
            parts.append(str(item))
        return "".join(parts).strip()
    return str(content)


def split_evenly(rows: list[RowPayload], group_count: int) -> list[list[RowPayload]]:
    group_count = max(1, min(group_count, len(rows)))
    chunk_size = (len(rows) + group_count - 1) // group_count
    return [rows[index:index + chunk_size] for index in range(0, len(rows), chunk_size)]


def append_error(state: RowState, message: str) -> None:
    text = str(message).strip()
    if not text:
        return
    if not state.error_message:
        state.error_message = text
        return
    existing = {part.strip() for part in state.error_message.split("；") if part.strip()}
    if text not in existing:
        state.error_message = f"{state.error_message}；{text}"


def build_failure_result(rows: list[RowPayload], error_message: str) -> dict[int, tuple[str, str, str]]:
    return {
        row.row_number: (PASS_RESULT_FAILED, "", error_message)
        for row in rows
    }


def validate_pass_rows(rows: list[RowPayload], result_rows: Any, batch_label: str) -> dict[int, tuple[str, str, str]]:
    if not isinstance(result_rows, list):
        raise ValueError(f"{batch_label} 返回结果不是 JSON 数组。")
    if len(result_rows) != len(rows):
        raise ValueError(f"{batch_label} 返回行数不一致：期望 {len(rows)}，实际 {len(result_rows)}。")

    validated: dict[int, tuple[str, str, str]] = {}
    for expected, actual in zip(rows, result_rows):
        if not isinstance(actual, dict):
            raise ValueError(f"{batch_label} 存在非对象结果：{actual}")
        row_number = actual.get("行号")
        result = normalize_header(actual.get("判定结果"))
        reason = normalize_header(actual.get("简短理由"))
        try:
            actual_row_number = int(row_number)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{batch_label} 行号非法：{actual}") from exc
        if actual_row_number != expected.row_number:
            raise ValueError(
                f"{batch_label} 行号顺序不一致：期望 {expected.row_number}，实际 {actual_row_number}。"
            )
        if result not in PASS_RESULT_MATCH:
            raise ValueError(f"{batch_label} 判定结果非法：{actual}")
        validated[expected.row_number] = (result, reason, "")
    return validated


def run_model_chunk(
    rows: list[RowPayload],
    rule_text: str,
    temperature: float,
    batch_label: str,
) -> dict[int, tuple[str, str, str]]:
    llm = build_chat_openai(
        temperature=temperature,
        max_tokens=8192,
        request_timeout=120,
        max_retries=3,
    )
    payload = [{"行号": row.row_number, "内容": row.content} for row in rows]
    response = llm.invoke(
        [
            SystemMessage(content=SYSTEM_PROMPT),
            HumanMessage(
                content=(
                    f"判定词/判定规则：{rule_text}\n\n"
                    f"待判定数据：\n{json.dumps(payload, ensure_ascii=False)}\n\n"
                    "请严格输出 JSON 数组。"
                )
            ),
        ]
    )
    response_text = extract_message_text(response)
    parsed = json.loads(clean_json_text(response_text))
    return validate_pass_rows(rows, parsed, batch_label)


def run_pass_batch(
    rows: list[RowPayload],
    rule_text: str,
    temperature: float,
    max_workers: int,
    batch_label: str,
) -> dict[int, tuple[str, str, str]]:
    if not rows:
        return {}

    worker_count = max(1, min(int(max_workers or 1), len(rows)))
    if worker_count == 1:
        try:
            return run_model_chunk(rows, rule_text, temperature, batch_label)
        except Exception as exc:
            return build_failure_result(rows, str(exc))

    chunks = split_evenly(rows, worker_count)
    indexed_results: dict[int, dict[int, tuple[str, str, str]]] = {}
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(run_model_chunk, chunk, rule_text, temperature, f"{batch_label}.{chunk_index}"): (chunk_index, chunk)
            for chunk_index, chunk in enumerate(chunks, start=1)
        }
        for future in as_completed(futures):
            chunk_index, chunk = futures[future]
            try:
                indexed_results[chunk_index] = future.result()
            except Exception as exc:
                indexed_results[chunk_index] = build_failure_result(chunk, str(exc))

    merged: dict[int, tuple[str, str, str]] = {}
    for chunk_index in sorted(indexed_results):
        merged.update(indexed_results[chunk_index])
    return merged


def write_output_headers(worksheet) -> dict[str, int]:
    start_column = worksheet.max_column + 1
    column_map: dict[str, int] = {}
    for offset, header in enumerate(OUTPUT_HEADERS):
        column_index = start_column + offset
        worksheet.cell(row=1, column=column_index, value=header)
        column_map[header] = column_index
    return column_map


def write_row_state(worksheet, column_map: dict[str, int], row_number: int, state: RowState) -> None:
    worksheet.cell(row=row_number, column=column_map["AI判定词"], value=state.rule_text)
    worksheet.cell(row=row_number, column=column_map["AI判定列"], value=state.target_columns_text)
    worksheet.cell(row=row_number, column=column_map["首轮判定"], value=state.pass1_result)
    worksheet.cell(row=row_number, column=column_map["首轮理由"], value=state.pass1_reason)
    worksheet.cell(row=row_number, column=column_map["复判判定"], value=state.pass2_result)
    worksheet.cell(row=row_number, column=column_map["复判理由"], value=state.pass2_reason)
    worksheet.cell(row=row_number, column=column_map["最终判定"], value=state.final_result)
    worksheet.cell(row=row_number, column=column_map["复核状态"], value=state.review_status)
    worksheet.cell(row=row_number, column=column_map["异常信息"], value=state.error_message)


def finalize_row_states(states: dict[int, RowState]) -> None:
    for state in states.values():
        if state.final_result == PASS_RESULT_EMPTY:
            state.review_status = REVIEW_STATUS_EMPTY
            continue

        if state.pass1_result in PASS_RESULT_MATCH and state.pass2_result in PASS_RESULT_MATCH:
            if state.pass1_result == state.pass2_result:
                state.final_result = state.pass1_result
                state.review_status = REVIEW_STATUS_CONSISTENT
            else:
                state.final_result = FINAL_RESULT_PENDING
                state.review_status = REVIEW_STATUS_PENDING
            continue

        if not state.final_result:
            state.final_result = FINAL_RESULT_PENDING
        if not state.review_status:
            state.review_status = REVIEW_STATUS_PENDING


def run_semantic_rule_judge(
    input_xlsx: str | Path,
    output_xlsx: str | Path,
    *,
    sheet_name: str,
    target_columns: list[str],
    rule_text: str,
    row_limit: int = 100,
    max_workers: int = 1,
    save_every_batches: int = 1,
    temperature: float = 0.1,
    sleep_seconds: float = 0.5,
    log_callback: Callable[[str], None] | None = None,
    stop_event=None,
    pause_event=None,
) -> str:
    input_path = Path(input_xlsx)
    output_path = Path(output_xlsx)
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    log = log_callback or print
    row_limit = max(1, int(row_limit or 1))
    max_workers = max(1, int(max_workers or 1))
    save_every_batches = max(1, int(save_every_batches or 1))
    rule_text = str(rule_text).strip()
    if not rule_text:
        raise ValueError("判定词/规则不能为空。")

    workbook = openpyxl.load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet 不存在：{sheet_name}")

    worksheet = workbook[sheet_name]
    headers = validate_headers([worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)])
    ordered_columns = resolve_target_columns(headers, target_columns)
    target_columns_text = "、".join(ordered_columns)
    column_map = write_output_headers(worksheet)

    row_payloads: list[RowPayload] = []
    row_states: dict[int, RowState] = {}
    for row_number in range(2, worksheet.max_row + 1):
        row_values = [worksheet.cell(row=row_number, column=index).value for index in range(1, len(headers) + 1)]
        content = build_row_content(headers, row_values, ordered_columns)
        row_payloads.append(RowPayload(row_number=row_number, content=content))
        state = RowState(rule_text=rule_text, target_columns_text=target_columns_text)
        if not content:
            state.pass1_result = PASS_RESULT_EMPTY
            state.pass2_result = PASS_RESULT_EMPTY
            state.final_result = PASS_RESULT_EMPTY
            state.review_status = REVIEW_STATUS_EMPTY
        row_states[row_number] = state
        write_row_state(worksheet, column_map, row_number, state)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    ai_rows = [row for row in row_payloads if row.content]
    total_batches = (len(ai_rows) + row_limit - 1) // row_limit if ai_rows else 0
    stopped_early = False

    def run_pass(pass_name: str, pass_index: int) -> None:
        nonlocal stopped_early
        batches_since_save = 0
        for batch_offset, start in enumerate(range(0, len(ai_rows), row_limit), start=1):
            if should_stop(stop_event):
                stopped_early = True
                log("任务已停止，保存当前结果。")
                break
            if wait_if_paused(pause_event, stop_event):
                stopped_early = True
                break

            batch_rows = ai_rows[start:start + row_limit]
            log(f"{pass_name}第 {batch_offset}/{total_batches} 批：{len(batch_rows)} 行")
            batch_results = run_pass_batch(
                batch_rows,
                rule_text,
                temperature,
                max_workers,
                f"{pass_name}{batch_offset}",
            )
            for row in batch_rows:
                result, reason, error_message = batch_results.get(
                    row.row_number,
                    (PASS_RESULT_FAILED, "", f"{pass_name}缺少返回结果"),
                )
                state = row_states[row.row_number]
                if pass_index == 1:
                    state.pass1_result = result
                    state.pass1_reason = reason
                else:
                    state.pass2_result = result
                    state.pass2_reason = reason
                if error_message:
                    append_error(state, error_message)
                write_row_state(worksheet, column_map, row.row_number, state)

            batches_since_save += 1
            if batches_since_save >= save_every_batches:
                workbook.save(output_path)
                batches_since_save = 0

            if sleep_seconds > 0:
                interruptible_sleep(sleep_seconds, stop_event)

        workbook.save(output_path)

    log(f"加载 Excel：{input_path}")
    log(f"目标 Sheet：{sheet_name}")
    log(f"判定列：{target_columns_text}")
    log(f"判定词/规则：{rule_text}")
    log(f"待判定行数：{len(ai_rows)}")

    run_pass("首轮", 1)
    if not stopped_early:
        run_pass("复判", 2)

    if stopped_early:
        for row in ai_rows:
            state = row_states[row.row_number]
            if not state.pass1_result:
                state.pass1_result = PASS_RESULT_FAILED
                append_error(state, "任务已停止，首轮未完成。")
            if not state.pass2_result:
                state.pass2_result = PASS_RESULT_FAILED
                append_error(state, "任务已停止，复判未完成。")

    finalize_row_states(row_states)
    for row_number, state in row_states.items():
        write_row_state(worksheet, column_map, row_number, state)
    workbook.save(output_path)
    workbook.close()
    return str(output_path)
