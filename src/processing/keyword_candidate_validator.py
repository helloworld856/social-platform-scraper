from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import openpyxl
from langchain_core.messages import HumanMessage, SystemMessage

from src.core import build_chat_openai, interruptible_sleep, should_stop, wait_if_paused
from src.processing.ai_semantic_rule_judge import (
    build_row_content,
    clean_json_text,
    extract_message_text,
    resolve_target_columns,
    validate_headers,
)
from src.tools.calibration import (
    KEYWORD_ROLE_CANDIDATE,
    KEYWORD_ROLE_OFFICIAL,
    STANDARDIZED_HEADERS,
)

AI_DETAIL_SHEET_NAME = "AI清洗明细"
SCORE_SHEET_NAME = "候选词评分"
COMPARE_SHEET_NAME = "跨游戏对比"

AI_DETAIL_HEADERS = [
    *STANDARDIZED_HEADERS,
    "官方关键词集",
    "AI判定列",
    "首轮判定",
    "首轮理由",
    "复判判定",
    "复判理由",
    "最终判定",
    "复核状态",
    "异常信息",
    "源Sheet",
    "源行号",
]

SCORE_HEADERS = [
    "游戏名",
    "平台",
    "语言",
    "官方关键词集",
    "候选词",
    "官方样本数",
    "候选样本数",
    "重合数",
    "新增数",
    "重合度",
    "新增占比",
    "候选纯度",
    "候选体量比",
    "待复核数",
    "校准值",
    "结论",
]

PASS_RESULT_MATCH = {"相关", "不相关"}
PASS_RESULT_EMPTY = "空内容"
PASS_RESULT_FAILED = "失败"
FINAL_RESULT_PENDING = "待人工复核"
REVIEW_STATUS_CONSISTENT = "一致"
REVIEW_STATUS_PENDING = "待人工复核"
REVIEW_STATUS_EMPTY = "空内容"

SYSTEM_PROMPT = """
你是一个游戏搜索结果清洗助手。
输入是一组待判断内容。每条内容都附带目标游戏名、平台、语言、官方关键词集合和原始文本。
你的任务是判断这条内容是否明确属于这个目标游戏或该游戏语境下的搜索结果。

判定要求：
1. “判定结果”只能是“相关”或“不相关”。
2. 重点排除同名异物、泛词误命中、借题发挥、蹭词内容。
3. 只根据提供的文本判断，不要猜测未提供的信息。
4. “简短理由”必须是一句短句。
5. 输出必须是标准 JSON 数组，不要输出 Markdown，不要输出代码块。

返回格式：
[
  {"行号": 2, "判定结果": "相关", "简短理由": "明确提到目标游戏内容"},
  {"行号": 3, "判定结果": "不相关", "简短理由": "内容不指向目标游戏"}
]
"""


@dataclass(frozen=True)
class ValidatorRowPayload:
    row_number: int
    game_name: str
    platform: str
    language: str
    keyword_role: str
    keyword_text: str
    official_keywords: tuple[str, ...]
    content: str


@dataclass
class ValidatorRowState:
    detail_row_number: int
    source_row_number: int
    source_sheet_name: str
    game_name: str
    platform: str
    language: str
    keyword_role: str
    keyword_text: str
    content_id: str
    content_link: str
    title: str
    description: str
    published_at: str
    source_output_file: str
    collection_status: str
    official_keywords_text: str
    target_columns_text: str
    pass1_result: str = ""
    pass1_reason: str = ""
    pass2_result: str = ""
    pass2_reason: str = ""
    final_result: str = ""
    review_status: str = ""
    error_message: str = ""


def validate_collection_headers(headers: list[str]) -> None:
    missing = [header for header in STANDARDIZED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"目标 Sheet 缺少标准化采集列：{', '.join(missing)}")


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_evenly(rows: list[ValidatorRowPayload], group_count: int) -> list[list[ValidatorRowPayload]]:
    group_count = max(1, min(group_count, len(rows)))
    chunk_size = (len(rows) + group_count - 1) // group_count
    return [rows[index:index + chunk_size] for index in range(0, len(rows), chunk_size)]


def append_error(state: ValidatorRowState, message: str) -> None:
    text = str(message).strip()
    if not text:
        return
    if not state.error_message:
        state.error_message = text
        return
    existing = {part.strip() for part in state.error_message.split("；") if part.strip()}
    if text not in existing:
        state.error_message = f"{state.error_message}；{text}"


def build_failure_result(rows: list[ValidatorRowPayload], error_message: str) -> dict[int, tuple[str, str, str]]:
    return {row.row_number: (PASS_RESULT_FAILED, "", error_message) for row in rows}


def validate_pass_rows(
    rows: list[ValidatorRowPayload],
    result_rows: Any,
    batch_label: str,
) -> dict[int, tuple[str, str, str]]:
    if not isinstance(result_rows, list):
        raise ValueError(f"{batch_label} 返回结果不是 JSON 数组。")
    if len(result_rows) != len(rows):
        raise ValueError(f"{batch_label} 返回行数不一致：期望 {len(rows)}，实际 {len(result_rows)}。")

    validated: dict[int, tuple[str, str, str]] = {}
    for expected, actual in zip(rows, result_rows):
        if not isinstance(actual, dict):
            raise ValueError(f"{batch_label} 存在非对象结果：{actual}")
        row_number = actual.get("行号")
        result = stringify_cell(actual.get("判定结果"))
        reason = stringify_cell(actual.get("简短理由"))
        try:
            actual_row_number = int(row_number)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{batch_label} 行号非法：{actual}") from exc
        if actual_row_number != expected.row_number:
            raise ValueError(f"{batch_label} 行号顺序不一致：期望 {expected.row_number}，实际 {actual_row_number}。")
        if result not in PASS_RESULT_MATCH:
            raise ValueError(f"{batch_label} 判定结果非法：{actual}")
        validated[expected.row_number] = (result, reason, "")
    return validated


def run_model_chunk(
    rows: list[ValidatorRowPayload],
    temperature: float,
    batch_label: str,
) -> dict[int, tuple[str, str, str]]:
    llm = build_chat_openai(
        temperature=temperature,
        max_tokens=8192,
        request_timeout=120,
        max_retries=3,
    )
    payload = [
        {
            "行号": row.row_number,
            "游戏名": row.game_name,
            "平台": row.platform,
            "语言": row.language,
            "关键词角色": row.keyword_role,
            "当前关键词": row.keyword_text,
            "官方关键词": list(row.official_keywords),
            "内容": row.content,
        }
        for row in rows
    ]
    response = llm.invoke(
        [
            SystemMessage(content=SYSTEM_PROMPT),
            HumanMessage(
                content=(
                    "请逐条判断这些内容是否属于对应的目标游戏搜索结果。\n"
                    "必须排除同名异物、泛词误命中、借题发挥内容。\n\n"
                    f"待判定数据：\n{json.dumps(payload, ensure_ascii=False)}\n\n"
                    "请严格输出 JSON 数组。"
                )
            ),
        ]
    )
    response_text = extract_message_text(response)
    parsed = json.loads(clean_json_text(response_text))
    return validate_pass_rows(rows, parsed, batch_label)


def run_validation_pass_batch(
    rows: list[ValidatorRowPayload],
    temperature: float,
    max_workers: int,
    batch_label: str,
) -> dict[int, tuple[str, str, str]]:
    if not rows:
        return {}

    worker_count = max(1, min(int(max_workers or 1), len(rows)))
    if worker_count == 1:
        try:
            return run_model_chunk(rows, temperature, batch_label)
        except Exception as exc:
            return build_failure_result(rows, str(exc))

    chunks = split_evenly(rows, worker_count)
    indexed_results: dict[int, dict[int, tuple[str, str, str]]] = {}
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(run_model_chunk, chunk, temperature, f"{batch_label}.{chunk_index}"): (chunk_index, chunk)
            for chunk_index, chunk in enumerate(chunks, start=1)
        }
        for future in as_completed(futures):
            chunk_index, chunk = futures[future]
            try:
                indexed_results[chunk_index] = future.result()
            except Exception as exc:
                indexed_results[chunk_index] = build_failure_result(chunk, str(exc))

    merged: dict[int, tuple[str, str, str]] = {}
    for chunk_index in sorted(indexed_results):
        merged.update(indexed_results[chunk_index])
    return merged


def write_ai_detail_headers(worksheet) -> None:
    worksheet.append(AI_DETAIL_HEADERS)
    worksheet.freeze_panes = "A2"


def write_score_headers(worksheet) -> None:
    worksheet.append(SCORE_HEADERS)
    worksheet.freeze_panes = "A2"


def write_ai_detail_row(worksheet, state: ValidatorRowState) -> None:
    worksheet.cell(row=state.detail_row_number, column=1, value=state.game_name)
    worksheet.cell(row=state.detail_row_number, column=2, value=state.platform)
    worksheet.cell(row=state.detail_row_number, column=3, value=state.language)
    worksheet.cell(row=state.detail_row_number, column=4, value=state.keyword_role)
    worksheet.cell(row=state.detail_row_number, column=5, value=state.keyword_text)
    worksheet.cell(row=state.detail_row_number, column=6, value=state.content_id)
    worksheet.cell(row=state.detail_row_number, column=7, value=state.content_link)
    worksheet.cell(row=state.detail_row_number, column=8, value=state.title)
    worksheet.cell(row=state.detail_row_number, column=9, value=state.description)
    worksheet.cell(row=state.detail_row_number, column=10, value=state.published_at)
    worksheet.cell(row=state.detail_row_number, column=11, value=state.source_output_file)
    worksheet.cell(row=state.detail_row_number, column=12, value=state.collection_status)
    worksheet.cell(row=state.detail_row_number, column=13, value=state.official_keywords_text)
    worksheet.cell(row=state.detail_row_number, column=14, value=state.target_columns_text)
    worksheet.cell(row=state.detail_row_number, column=15, value=state.pass1_result)
    worksheet.cell(row=state.detail_row_number, column=16, value=state.pass1_reason)
    worksheet.cell(row=state.detail_row_number, column=17, value=state.pass2_result)
    worksheet.cell(row=state.detail_row_number, column=18, value=state.pass2_reason)
    worksheet.cell(row=state.detail_row_number, column=19, value=state.final_result)
    worksheet.cell(row=state.detail_row_number, column=20, value=state.review_status)
    worksheet.cell(row=state.detail_row_number, column=21, value=state.error_message)
    worksheet.cell(row=state.detail_row_number, column=22, value=state.source_sheet_name)
    worksheet.cell(row=state.detail_row_number, column=23, value=state.source_row_number)


def finalize_row_states(states: dict[int, ValidatorRowState]) -> None:
    for state in states.values():
        if state.final_result == PASS_RESULT_EMPTY:
            state.review_status = REVIEW_STATUS_EMPTY
            continue

        if state.pass1_result in PASS_RESULT_MATCH and state.pass2_result in PASS_RESULT_MATCH:
            if state.pass1_result == state.pass2_result:
                state.final_result = state.pass1_result
                state.review_status = REVIEW_STATUS_CONSISTENT
            else:
                state.final_result = FINAL_RESULT_PENDING
                state.review_status = REVIEW_STATUS_PENDING
            continue

        if not state.final_result:
            state.final_result = FINAL_RESULT_PENDING
        if not state.review_status:
            state.review_status = REVIEW_STATUS_PENDING


def unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = stringify_cell(value)
        lowered = text.casefold()
        if not text or lowered in seen:
            continue
        seen.add(lowered)
        result.append(text)
    return result


def safe_ratio(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return numerator / denominator


def round_ratio(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 4)


def round_score(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 2)


def build_score_rows(states: dict[int, ValidatorRowState]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for state in sorted(states.values(), key=lambda item: item.detail_row_number):
        group_key = (state.game_name, state.platform, state.language)
        bucket = groups.setdefault(
            group_key,
            {
                "official_keywords": unique_preserve_order(state.official_keywords_text.split("、")) if state.official_keywords_text else [],
                "official_rows": [],
                "candidate_order": [],
                "candidate_rows": {},
            },
        )
        if state.keyword_role == KEYWORD_ROLE_OFFICIAL:
            bucket["official_rows"].append(state)
            if not bucket["official_keywords"] and state.keyword_text:
                bucket["official_keywords"] = unique_preserve_order(bucket["official_keywords"] + [state.keyword_text])
        elif state.keyword_role == KEYWORD_ROLE_CANDIDATE:
            if state.keyword_text not in bucket["candidate_rows"]:
                bucket["candidate_rows"][state.keyword_text] = []
                bucket["candidate_order"].append(state.keyword_text)
            bucket["candidate_rows"][state.keyword_text].append(state)

    score_rows: list[dict[str, Any]] = []
    for (game_name, platform, language), bucket in groups.items():
        official_rows: list[ValidatorRowState] = bucket["official_rows"]
        official_keywords = unique_preserve_order(bucket["official_keywords"])
        official_clean_ids = {
            row.content_id
            for row in official_rows
            if row.final_result == "相关" and row.content_id
        }
        official_pending_rows = sum(1 for row in official_rows if row.review_status == REVIEW_STATUS_PENDING)
        total_official_rows = len(official_rows)

        for candidate_keyword in bucket["candidate_order"]:
            candidate_rows: list[ValidatorRowState] = bucket["candidate_rows"].get(candidate_keyword, [])
            candidate_clean_ids = {
                row.content_id
                for row in candidate_rows
                if row.final_result == "相关" and row.content_id
            }
            candidate_pending_rows = sum(1 for row in candidate_rows if row.review_status == REVIEW_STATUS_PENDING)
            total_candidate_rows = len(candidate_rows)

            official_count = len(official_clean_ids)
            candidate_count = len(candidate_clean_ids)
            overlap_count = len(official_clean_ids & candidate_clean_ids)
            incremental_count = len(candidate_clean_ids - official_clean_ids)

            overlap_ratio = safe_ratio(overlap_count, official_count)
            incremental_ratio = safe_ratio(incremental_count, official_count)
            candidate_purity = safe_ratio(overlap_count, candidate_count)
            candidate_volume_ratio = safe_ratio(candidate_count, official_count)
            review_pending_count = official_pending_rows + candidate_pending_rows
            review_pending_ratio = safe_ratio(review_pending_count, total_official_rows + total_candidate_rows) or 0.0

            overlap_balance = 0.0
            incremental_score = 0.0
            confidence = 0.0
            calibration_score: float | None = None

            if official_count > 0 and candidate_count > 0 and overlap_ratio is not None and incremental_ratio is not None and candidate_purity is not None:
                overlap_balance = max(0.0, 1 - abs(overlap_ratio - 0.4) / 0.4)
                incremental_score = min(1.0, incremental_ratio / 0.3)
                confidence = min(1.0, official_count / 30, candidate_count / 15) * max(0.0, 1 - review_pending_ratio)
                calibration_score = 100 * (
                    0.45 * overlap_balance
                    + 0.30 * candidate_purity
                    + 0.15 * incremental_score
                    + 0.10 * confidence
                )

            if official_count == 0 or candidate_count == 0:
                conclusion = "无有效样本"
            elif (candidate_purity or 0.0) < 0.15:
                conclusion = "噪声过高"
            elif confidence < 0.4:
                conclusion = "样本不足"
            elif (calibration_score or 0.0) >= 75:
                conclusion = "推荐"
            elif (calibration_score or 0.0) >= 55:
                conclusion = "可观察"
            else:
                conclusion = "不推荐"

            score_rows.append(
                {
                    "游戏名": game_name,
                    "平台": platform,
                    "语言": language,
                    "官方关键词集": "、".join(official_keywords),
                    "候选词": candidate_keyword,
                    "官方样本数": official_count,
                    "候选样本数": candidate_count,
                    "重合数": overlap_count,
                    "新增数": incremental_count,
                    "重合度": round_ratio(overlap_ratio),
                    "新增占比": round_ratio(incremental_ratio),
                    "候选纯度": round_ratio(candidate_purity),
                    "候选体量比": round_ratio(candidate_volume_ratio),
                    "待复核数": review_pending_count,
                    "校准值": round_score(calibration_score),
                    "结论": conclusion,
                }
            )
    return score_rows


def write_score_rows(worksheet, score_rows: list[dict[str, Any]]) -> None:
    for row in score_rows:
        worksheet.append([row.get(header, "") for header in SCORE_HEADERS])


def sort_compare_rows(score_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def sort_key(item: dict[str, Any]) -> tuple[float, str, str, str, str]:
        score = item.get("校准值")
        return (float(score) if score is not None else -1.0, item["候选词"], item["游戏名"], item["平台"], item["语言"])

    return sorted(score_rows, key=sort_key, reverse=True)


def remove_existing_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        workbook.remove(worksheet)


def run_keyword_candidate_validator(
    input_xlsx: str | Path,
    output_xlsx: str | Path,
    *,
    sheet_name: str,
    target_columns: list[str],
    row_limit: int = 100,
    max_workers: int = 1,
    save_every_batches: int = 1,
    temperature: float = 0.1,
    sleep_seconds: float = 0.5,
    log_callback: Callable[[str], None] | None = None,
    stop_event=None,
    pause_event=None,
) -> str:
    input_path = Path(input_xlsx)
    output_path = Path(output_xlsx)
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    log = log_callback or print
    row_limit = max(1, int(row_limit or 1))
    max_workers = max(1, int(max_workers or 1))
    save_every_batches = max(1, int(save_every_batches or 1))

    workbook = openpyxl.load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet 不存在：{sheet_name}")

    source_sheet = workbook[sheet_name]
    headers = validate_headers([source_sheet.cell(row=1, column=index).value for index in range(1, source_sheet.max_column + 1)])
    validate_collection_headers(headers)
    ordered_columns = resolve_target_columns(headers, target_columns)
    target_columns_text = "、".join(ordered_columns)
    header_index = {header: index + 1 for index, header in enumerate(headers)}

    remove_existing_sheet(workbook, AI_DETAIL_SHEET_NAME)
    remove_existing_sheet(workbook, SCORE_SHEET_NAME)
    remove_existing_sheet(workbook, COMPARE_SHEET_NAME)
    detail_sheet = workbook.create_sheet(AI_DETAIL_SHEET_NAME)
    score_sheet = workbook.create_sheet(SCORE_SHEET_NAME)
    compare_sheet = workbook.create_sheet(COMPARE_SHEET_NAME)
    write_ai_detail_headers(detail_sheet)
    write_score_headers(score_sheet)
    write_score_headers(compare_sheet)

    raw_rows: list[tuple[int, dict[str, str], list[Any]]] = []
    for row_number in range(2, source_sheet.max_row + 1):
        row_values = [source_sheet.cell(row=row_number, column=index).value for index in range(1, len(headers) + 1)]
        row_dict = {header: stringify_cell(row_values[index]) for index, header in enumerate(headers)}
        raw_rows.append((row_number, row_dict, row_values))

    group_official_keywords: dict[tuple[str, str, str], list[str]] = {}
    for _, row_dict, _ in raw_rows:
        role = row_dict["关键词角色"]
        group_key = (row_dict["游戏名"], row_dict["平台"], row_dict["语言"])
        if role != KEYWORD_ROLE_OFFICIAL:
            continue
        keywords = group_official_keywords.setdefault(group_key, [])
        keyword_text = row_dict["关键词文本"]
        if keyword_text and keyword_text.casefold() not in {item.casefold() for item in keywords}:
            keywords.append(keyword_text)

    row_payloads: list[ValidatorRowPayload] = []
    row_states: dict[int, ValidatorRowState] = {}
    detail_row_number = 2
    for source_row_number, row_dict, row_values in raw_rows:
        role = row_dict["关键词角色"]
        if role not in {KEYWORD_ROLE_OFFICIAL, KEYWORD_ROLE_CANDIDATE}:
            raise ValueError(f"发现非法关键词角色：{role or '<empty>'}（源行 {source_row_number}）")

        game_name = row_dict["游戏名"]
        platform = row_dict["平台"]
        language = row_dict["语言"]
        if not game_name or not platform or not language:
            raise ValueError(f"标准化采集行缺少游戏名/平台/语言（源行 {source_row_number}）")

        group_key = (game_name, platform, language)
        official_keywords = tuple(group_official_keywords.get(group_key, []))
        content = build_row_content(headers, row_values, ordered_columns)
        state = ValidatorRowState(
            detail_row_number=detail_row_number,
            source_row_number=source_row_number,
            source_sheet_name=sheet_name,
            game_name=game_name,
            platform=platform,
            language=language,
            keyword_role=role,
            keyword_text=row_dict["关键词文本"],
            content_id=row_dict["内容ID"],
            content_link=row_dict["内容链接"],
            title=row_dict["标题"],
            description=row_dict["简介"],
            published_at=row_dict["发布时间"],
            source_output_file=row_dict["原始输出文件"],
            collection_status=row_dict["采集状态"],
            official_keywords_text="、".join(official_keywords),
            target_columns_text=target_columns_text,
        )
        if not content:
            state.pass1_result = PASS_RESULT_EMPTY
            state.pass2_result = PASS_RESULT_EMPTY
            state.final_result = PASS_RESULT_EMPTY
            state.review_status = REVIEW_STATUS_EMPTY
        else:
            row_payloads.append(
                ValidatorRowPayload(
                    row_number=source_row_number,
                    game_name=game_name,
                    platform=platform,
                    language=language,
                    keyword_role=role,
                    keyword_text=row_dict["关键词文本"],
                    official_keywords=official_keywords,
                    content=content,
                )
            )
        row_states[source_row_number] = state
        write_ai_detail_row(detail_sheet, state)
        detail_row_number += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    ai_rows = [row for row in row_payloads if row.content]
    total_batches = (len(ai_rows) + row_limit - 1) // row_limit if ai_rows else 0
    stopped_early = False

    def run_pass(pass_name: str, pass_index: int) -> None:
        nonlocal stopped_early
        batches_since_save = 0
        for batch_offset, start in enumerate(range(0, len(ai_rows), row_limit), start=1):
            if should_stop(stop_event):
                stopped_early = True
                log("任务已停止，保存当前结果。")
                break
            if wait_if_paused(pause_event, stop_event):
                stopped_early = True
                break

            batch_rows = ai_rows[start:start + row_limit]
            log(f"{pass_name}第 {batch_offset}/{total_batches} 批：{len(batch_rows)} 行")
            batch_results = run_validation_pass_batch(
                batch_rows,
                temperature,
                max_workers,
                f"{pass_name}{batch_offset}",
            )
            for row in batch_rows:
                result, reason, error_message = batch_results.get(
                    row.row_number,
                    (PASS_RESULT_FAILED, "", f"{pass_name}缺少返回结果"),
                )
                state = row_states[row.row_number]
                if pass_index == 1:
                    state.pass1_result = result
                    state.pass1_reason = reason
                else:
                    state.pass2_result = result
                    state.pass2_reason = reason
                if error_message:
                    append_error(state, error_message)
                write_ai_detail_row(detail_sheet, state)

            batches_since_save += 1
            if batches_since_save >= save_every_batches:
                workbook.save(output_path)
                batches_since_save = 0

            if sleep_seconds > 0:
                interruptible_sleep(sleep_seconds, stop_event)

        workbook.save(output_path)

    log(f"加载 Excel：{input_path}")
    log(f"目标 Sheet：{sheet_name}")
    log(f"判定列：{target_columns_text}")
    log(f"待清洗行数：{len(ai_rows)}")

    run_pass("首轮", 1)
    if not stopped_early:
        run_pass("复判", 2)

    if stopped_early:
        for row in ai_rows:
            state = row_states[row.row_number]
            if not state.pass1_result:
                state.pass1_result = PASS_RESULT_FAILED
                append_error(state, "任务已停止，首轮未完成。")
            if not state.pass2_result:
                state.pass2_result = PASS_RESULT_FAILED
                append_error(state, "任务已停止，复判未完成。")

    finalize_row_states(row_states)
    for state in row_states.values():
        write_ai_detail_row(detail_sheet, state)

    score_rows = build_score_rows(row_states)
    write_score_rows(score_sheet, score_rows)
    write_score_rows(compare_sheet, sort_compare_rows(score_rows))

    workbook.save(output_path)
    workbook.close()
    return str(output_path)
