# -*- coding: utf-8 -*-
"""Excel 数据文件合并与规范化清洗模块。

本模块提供将多个零散的同结构 Excel 数据文件合并为单个文件的工作，
支持依据关键字过滤文件名、统一“序号”字段对齐行号，并实现数据清洗防护。
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.core import (
    ArtifactRef,
    RunError,
    RunOutcome,
    RunStatus,
    build_run_output_dir,
    generate_run_id,
    sanitize_xlsx_cell,
    should_stop,
    wait_if_paused,
)

# 平台别名至归一化前缀的映射表
PLATFORM_PREFIX = {
    "youtube": "youtube",
    "tiktok": "tiktok",
    "x": "x",
    "twitter": "x",
}


def normalize_platform(platform: str) -> str:
    """归一化平台标识名称。"""
    value = (platform or "").strip().lower()
    if value in PLATFORM_PREFIX:
        return PLATFORM_PREFIX[value]
    return value or "merged"


def find_xlsx_files(folder: str | Path, keyword: str, output_file: str | Path | None = None) -> list[Path]:
    """扫描指定目录下符合名称关键字条件的所有 Excel (.xlsx) 文件，并过滤掉临时文件及输出目标自身。"""
    folder_path = Path(folder)
    keyword = (keyword or "").strip().lower()
    output_name = Path(output_file).name.lower() if output_file else ""
    files: list[Path] = []
    for path in sorted(folder_path.glob("*.xlsx")):
        name = path.name.lower()
        if output_name and name == output_name:
            continue
        if path.name.startswith("~$"):
            continue
        if keyword and keyword not in name:
            continue
        files.append(path)
    return files


def _normalize_headers(raw_headers) -> list[str]:
    """表头行数据清洗与规范化，去除两端空白，丢弃尾部空单元格。"""
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def merge_xlsx_files(
    folder: str | Path,
    keyword: str = "keyword",
    platform: str = "merged",
    output_file: str | Path | None = None,
    schema_mode: str = "union_schema",  # "strict" or "union_schema"
    run_id: str | None = None,
    stop_event = None,
    pause_event = None,
) -> RunOutcome:
    """合并指定文件夹下的多个 Excel 文件，支持 strict 模式和 union_schema 并集模式。

    每次执行在 output/xlsx_merge/<run_id>/ 目录下生成独立产物与 merge_report.json 报告。

    Args:
        folder: 源 Excel 文件所在的目录。
        keyword: 文件名包含的关键字，用于定位要合并的子集。
        platform: 对应的平台前缀，作为自动命名时子文件夹与前缀标识。
        output_file: 可选。指定合并后导出的文件名或完整绝对路径。
        schema_mode: 合并模式，"strict" 或 "union_schema"。
        run_id: 任务执行 ID，留空则自动生成。
        stop_event: 停止事件信号。
        pause_event: 暂停事件信号。

    Returns:
        RunOutcome: 标准化执行结果对象。
    """
    platform_prefix = normalize_platform(platform)
    actual_run_id = run_id or generate_run_id()
    run_dir = build_run_output_dir("xlsx_merge", actual_run_id)

    # 确定输出文件名与路径
    if output_file:
        dest_name = Path(output_file).name
    else:
        dest_name = f"{platform_prefix}_merge.xlsx"
    final_xlsx_path = run_dir / dest_name
    final_report_path = run_dir / "merge_report.json"

    outcome = RunOutcome(
        run_id=actual_run_id,
        tool_id="xlsx_merge",
        status=RunStatus.SUCCEEDED,
    )

    files = find_xlsx_files(folder, keyword, final_xlsx_path)
    outcome.stats.input_count = len(files)

    if not files:
        outcome.status = RunStatus.FAILED
        outcome.errors.append(
            RunError(
                code="XLSX_FILE_UNREADABLE",
                message=f"在目录 {folder} 中未找到符合关键字 “{keyword}” 的合并目标文件",
            )
        )
        outcome.save_to_json(final_report_path)
        return outcome

    # ==========================================
    # Pass 1: 扫描各个文件，校验表头 schema
    # ==========================================
    valid_sheets: list[tuple[Path, str, list[str]]] = []
    base_headers: list[str] | None = None
    union_headers: list[str] = []
    union_set: set[str] = set()

    for file_path in files:
        if should_stop(stop_event):
            break

        wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                row_iter = ws.iter_rows(values_only=True)
                try:
                    raw_headers = next(row_iter, [])
                except StopIteration:
                    raw_headers = []
                source_headers = _normalize_headers(raw_headers)

                # 空工作表过滤
                if not source_headers or all(not value for value in source_headers):
                    outcome.errors.append(
                        RunError(
                            code="XLSX_WORKSHEET_EMPTY",
                            message=f"工作表 {ws.title} 没有有效表头列，跳过该页",
                            item=str(file_path),
                        )
                    )
                    outcome.stats.skipped_count += 1
                    continue

                # 重复表头过滤
                if len(source_headers) != len(set(source_headers)):
                    outcome.errors.append(
                        RunError(
                            code="XLSX_DUPLICATE_HEADER",
                            message=f"工作表 {ws.title} 包含重复的表头列：{source_headers}，跳过该页",
                            item=str(file_path),
                        )
                    )
                    outcome.stats.failed_count += 1
                    continue

                # 初始化基础表头
                if base_headers is None:
                    base_headers = list(source_headers)
                    if "序号" in base_headers:
                        # 确保“序号”始终放在第一列
                        union_headers = ["序号"] + [h for h in base_headers if h != "序号"]
                    else:
                        union_headers = ["序号"] + base_headers
                    union_set = set(union_headers)

                # Schema 匹配逻辑
                if schema_mode == "strict":
                    # 严格模式：要求列名和顺序完全一致
                    if source_headers != base_headers:
                        outcome.errors.append(
                            RunError(
                                code="XLSX_SCHEMA_MISMATCH",
                                message=f"工作表 {ws.title} 表头结构不一致，拒绝合并。预期：{base_headers}，实际：{source_headers}",
                                item=str(file_path),
                            )
                        )
                        outcome.stats.failed_count += 1
                        continue
                else:
                    # 并集模式：收集新增列
                    for header in source_headers:
                        if header != "序号" and header not in union_set:
                            union_headers.append(header)
                            union_set.add(header)

                valid_sheets.append((file_path, ws.title, source_headers))

        except Exception as exc:
            outcome.errors.append(
                RunError(
                    code="XLSX_FILE_UNREADABLE",
                    message=f"文件损坏或不可读：{exc}",
                    item=str(file_path),
                )
            )
            outcome.stats.failed_count += 1
        finally:
            if wb is not None:
                wb.close()

    # 中途取消检测
    if should_stop(stop_event):
        outcome.status = RunStatus.CANCELLED
        outcome.errors.append(RunError(code="RUN_CANCELLED", message="合并任务已被用户停止"))
        outcome.save_to_json(final_report_path)
        return outcome

    # 如果没有任何有效的数据页，标记失败
    if not valid_sheets:
        outcome.status = RunStatus.FAILED
        outcome.errors.append(RunError(code="XLSX_WORKSHEET_EMPTY", message="没有找到任何包含有效表头的 Excel 数据页"))
        outcome.save_to_json(final_report_path)
        return outcome

    # ==========================================
    # Pass 2: 初始化最终 Workbook 并合并数据行
    # ==========================================
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "合并数据"
    output_ws.append(union_headers)

    serial_col_index = union_headers.index("序号")
    current_no = 1
    total_merged_rows = 0

    for file_path, sheet_title, source_headers in valid_sheets:
        if should_stop(stop_event):
            break
        wait_if_paused(pause_event, stop_event)

        wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_title]
            row_iter = ws.iter_rows(values_only=True)
            next(row_iter, None)  # 掠过首行表头

            source_index = {name: idx for idx, name in enumerate(source_headers)}

            for row_values in row_iter:
                # 过滤全空行
                if not row_values or all(value is None or str(value).strip() == "" for value in row_values):
                    continue

                output_row = []
                for column_index, header in enumerate(union_headers):
                    if column_index == serial_col_index:
                        output_row.append(current_no)
                    else:
                        source_pos = source_index.get(header)
                        # 如果当前子表缺失该并集字段，自动填补空字串
                        val = row_values[source_pos] if source_pos is not None and source_pos < len(row_values) else ""
                        output_row.append(sanitize_xlsx_cell(val))

                output_ws.append(output_row)
                current_no += 1
                total_merged_rows += 1

            outcome.stats.success_count += 1
        except Exception as exc:
            outcome.errors.append(
                RunError(
                    code="XLSX_FILE_UNREADABLE",
                    message=f"数据读取中途发生未知异常：{exc}",
                    item=str(file_path),
                )
            )
            outcome.stats.failed_count += 1
        finally:
            if wb is not None:
                wb.close()

    # 再次检查中途取消
    if should_stop(stop_event):
        outcome.status = RunStatus.CANCELLED
        outcome.errors.append(RunError(code="RUN_CANCELLED", message="合并任务在合并过程中被用户中止"))
        outcome.save_to_json(final_report_path)
        return outcome

    # 如果最后合并的有效行数为0，不保留产物并报错
    if total_merged_rows == 0:
        outcome.status = RunStatus.FAILED
        outcome.errors.append(RunError(code="XLSX_WORKSHEET_EMPTY", message="没有成功合并到任何有效数据行"))
        outcome.save_to_json(final_report_path)
        return outcome

    # 确定终态状态
    if outcome.stats.failed_count > 0 or outcome.stats.skipped_count > 0:
        outcome.status = RunStatus.PARTIAL
    else:
        outcome.status = RunStatus.SUCCEEDED

    # 写入 XLSX 文件保护 (临时文件保存后重命名原子替换)
    temp_fd = None
    temp_path = None
    try:
        import tempfile
        temp_fd, temp_path = tempfile.mkstemp(dir=str(final_xlsx_path.parent), suffix=".tmp")
        os.close(temp_fd)  # 立即关闭，以便 openpyxl 写入
        output_wb.save(temp_path)
        os.replace(temp_path, str(final_xlsx_path))
        outcome.output_path = str(final_xlsx_path)
        outcome.artifacts.append(
            ArtifactRef(path=str(final_xlsx_path), label="合并后的Excel数据")
        )
    except Exception as exc:
        outcome.status = RunStatus.FAILED
        outcome.errors.append(
            RunError(
                code="XLSX_FILE_UNREADABLE",
                message=f"保存合并后 Excel 临时文件失败：{exc}",
            )
        )
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception:
                pass
    
    outcome.save_to_json(final_report_path)
    outcome.artifacts.append(
        ArtifactRef(path=str(final_report_path), label="合并任务报告")
    )
    return outcome


def main(argv=None):
    """合并脚本入口函数，解析命令行参数并执行。"""
    parser = argparse.ArgumentParser(description="合并多个 xlsx 文件")
    parser.add_argument("folder", help="包含 xlsx 文件的文件夹")
    parser.add_argument("--keyword", default="keyword", help="文件名包含的关键词，留空则合并所有 xlsx")
    parser.add_argument("--platform", default="merged", help="平台前缀，例如 youtube/tiktok/x")
    parser.add_argument("--output", default="", help="输出 xlsx 路径，不填则自动写入 output/xlsx_merge/<run_id>")
    parser.add_argument("--schema_mode", default="union_schema", choices=["union_schema", "strict"], help="表头模式")
    args = parser.parse_args(argv)
    
    outcome = merge_xlsx_files(
        args.folder,
        args.keyword,
        args.platform,
        args.output or None,
        schema_mode=args.schema_mode
    )
    print(f"Status: {outcome.status.value}")
    if outcome.output_path:
        print(f"Merged output saved to: {outcome.output_path}")
    else:
        print("Merge failed or no output produced.")


if __name__ == "__main__":
    main()
