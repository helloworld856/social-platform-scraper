from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

pytest.importorskip("PyQt5")

from PyQt5.QtWidgets import QApplication

from src.processing.keyword_candidate_validator import (
    AI_DETAIL_SHEET_NAME,
    COMPARE_SHEET_NAME,
    SCORE_SHEET_NAME,
    run_keyword_candidate_validator,
)
from src.processing.windows import KeywordCandidateValidatorWindow
from src.tools.calibration import KEYWORD_ROLE_CANDIDATE, KEYWORD_ROLE_OFFICIAL, STANDARDIZED_HEADERS, STANDARDIZED_SHEET_NAME

_app = None


def _ensure_app():
    global _app
    if _app is None:
        _app = QApplication.instance() or QApplication([])
    return _app


def _append_standardized_row(sheet, **kwargs):
    row = [
        kwargs.get("游戏名", "Game A"),
        kwargs.get("平台", "youtube"),
        kwargs.get("语言", "en"),
        kwargs.get("关键词角色", KEYWORD_ROLE_OFFICIAL),
        kwargs.get("关键词文本", "base"),
        kwargs.get("内容ID", ""),
        kwargs.get("内容链接", ""),
        kwargs.get("标题", ""),
        kwargs.get("简介", ""),
        kwargs.get("发布时间", "2026-06-18"),
        kwargs.get("原始输出文件", "source.xlsx"),
        kwargs.get("采集状态", "SUCCESS"),
    ]
    sheet.append(row)


def _create_validator_workbook(path: Path) -> dict[str, tuple[str, str]]:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = STANDARDIZED_SHEET_NAME
    sheet.append(STANDARDIZED_HEADERS)

    pass_policy: dict[int, tuple[str, str]] = {}

    for index in range(1, 31):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_OFFICIAL,
            关键词文本="Game A",
            内容ID=f"o{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'a' * 9}{index:02d}"[-11:],
            标题=f"Game A official {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")

    for index in range(1, 13):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_CANDIDATE,
            关键词文本="alias_recommended",
            内容ID=f"o{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'b' * 9}{index:02d}"[-11:],
            标题=f"Game A alias recommended overlap {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")
    for index in range(1, 4):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_CANDIDATE,
            关键词文本="alias_recommended",
            内容ID=f"n{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'c' * 9}{index:02d}"[-11:],
            标题=f"Game A alias recommended new {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")

    for index in range(1, 10):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_CANDIDATE,
            关键词文本="alias_observable",
            内容ID=f"o{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'d' * 9}{index:02d}"[-11:],
            标题=f"Game A alias observable overlap {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")
    for index in range(1, 7):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_CANDIDATE,
            关键词文本="alias_observable",
            内容ID=f"m{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'e' * 9}{index:02d}"[-11:],
            标题=f"Game A alias observable new {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")

    for index in range(1, 16):
        _append_standardized_row(
            sheet,
            关键词角色=KEYWORD_ROLE_CANDIDATE,
            关键词文本="alias_noise",
            内容ID=f"z{index:02d}",
            内容链接=f"https://www.youtube.com/watch?v={'f' * 9}{index:02d}"[-11:],
            标题=f"Game A alias noise {index}",
        )
        pass_policy[sheet.max_row] = ("相关", "相关")

    _append_standardized_row(
        sheet,
        关键词角色=KEYWORD_ROLE_CANDIDATE,
        关键词文本="alias_small",
        内容ID="o01",
        内容链接="https://www.youtube.com/watch?v=smallsample",
        标题="Game A alias small sample",
    )
    pass_policy[sheet.max_row] = ("相关", "相关")

    _append_standardized_row(
        sheet,
        关键词角色=KEYWORD_ROLE_CANDIDATE,
        关键词文本="alias_pending",
        内容ID="o02",
        内容链接="https://www.youtube.com/watch?v=pendingrow11",
        标题="Game A alias pending",
    )
    pass_policy[sheet.max_row] = ("相关", "不相关")

    meta = workbook.create_sheet("Meta")
    meta.append(["保留"])
    meta.append(["value"])
    workbook.save(path)
    workbook.close()
    return pass_policy


def test_window_loads_sheets_and_defaults_columns(tmp_path):
    _ensure_app()
    file_path = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = STANDARDIZED_SHEET_NAME
    sheet.append(STANDARDIZED_HEADERS)
    workbook.create_sheet("Meta")
    workbook.save(file_path)
    workbook.close()

    window = KeywordCandidateValidatorWindow()
    window.widgets["input_xlsx"].path_edit.setText(str(file_path))
    window._reload_sheet_names(str(file_path))

    combo = window.widgets["sheet_name"]
    assert combo.count() == 2
    assert combo.currentText() == STANDARDIZED_SHEET_NAME
    checked = [window.column_list.item(i).text() for i in range(window.column_list.count()) if window.column_list.item(i).checkState() == 2]
    assert "标题" in checked
    assert "简介" in checked
    window.deleteLater()


def test_run_keyword_candidate_validator_outputs_detail_and_scores(tmp_path, monkeypatch):
    input_path = tmp_path / "validator_input.xlsx"
    output_path = tmp_path / "validator_output.xlsx"
    pass_policy = _create_validator_workbook(input_path)

    def fake_run_validation_pass_batch(rows, temperature, max_workers, batch_label):
        is_pass1 = batch_label.startswith("首轮")
        results = {}
        for row in rows:
            first, second = pass_policy[row.row_number]
            result = first if is_pass1 else second
            results[row.row_number] = (result, f"{batch_label}-{row.row_number}", "")
        return results

    monkeypatch.setattr("src.processing.keyword_candidate_validator.run_validation_pass_batch", fake_run_validation_pass_batch)

    result = run_keyword_candidate_validator(
        input_path,
        output_path,
        sheet_name=STANDARDIZED_SHEET_NAME,
        target_columns=["标题", "简介"],
        row_limit=20,
        max_workers=3,
        save_every_batches=1,
        temperature=0.1,
        sleep_seconds=0.0,
    )

    assert result == str(output_path)
    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == [STANDARDIZED_SHEET_NAME, "Meta", AI_DETAIL_SHEET_NAME, SCORE_SHEET_NAME, COMPARE_SHEET_NAME]

    detail_sheet = workbook[AI_DETAIL_SHEET_NAME]
    score_sheet = workbook[SCORE_SHEET_NAME]
    compare_sheet = workbook[COMPARE_SHEET_NAME]

    score_headers = [score_sheet.cell(row=1, column=index).value for index in range(1, score_sheet.max_column + 1)]
    score_rows = [
        {
            str(score_headers[index - 1]): score_sheet.cell(row=row_number, column=index).value
            for index in range(1, score_sheet.max_column + 1)
        }
        for row_number in range(2, score_sheet.max_row + 1)
    ]
    score_map = {row["候选词"]: row for row in score_rows}

    assert score_map["alias_recommended"]["官方样本数"] == 30
    assert score_map["alias_recommended"]["候选样本数"] == 15
    assert score_map["alias_recommended"]["重合数"] == 12
    assert score_map["alias_recommended"]["新增数"] == 3
    assert score_map["alias_recommended"]["校准值"] == 84
    assert score_map["alias_recommended"]["结论"] == "推荐"

    assert score_map["alias_observable"]["校准值"] == 71.75
    assert score_map["alias_observable"]["结论"] == "可观察"
    assert score_map["alias_noise"]["结论"] == "噪声过高"
    assert score_map["alias_small"]["结论"] == "样本不足"
    assert score_map["alias_pending"]["待复核数"] == 1
    assert score_map["alias_pending"]["结论"] == "无有效样本"

    assert compare_sheet.cell(row=2, column=5).value == "alias_recommended"

    pending_row = None
    for row_number in range(2, detail_sheet.max_row + 1):
        if detail_sheet.cell(row=row_number, column=5).value == "alias_pending":
            pending_row = row_number
            break
    assert pending_row is not None
    assert detail_sheet.cell(row=pending_row, column=19).value == "待人工复核"
    assert detail_sheet.cell(row=pending_row, column=20).value == "待人工复核"
    assert workbook["Meta"].cell(row=2, column=1).value == "value"
    workbook.close()
