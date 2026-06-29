# -*- coding: utf-8 -*-
import threading
from pathlib import Path
from openpyxl import Workbook
from src.processing.xlsx_merge import merge_xlsx_files, RunStatus

def create_xlsx_file(path: Path, headers: list[str], rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if headers:
        ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)

def test_merge_same_headers(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    f2 = tmp_path / "f2.xlsx"
    create_xlsx_file(f1, ["序号", "标题", "播放量"], [[1, "视频1", 100], [2, "视频2", 200]])
    create_xlsx_file(f2, ["序号", "标题", "播放量"], [[1, "视频3", 300]])
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run")
    assert outcome.status == RunStatus.SUCCEEDED
    assert outcome.stats.success_count == 2
    assert outcome.stats.failed_count == 0
    assert outcome.stats.skipped_count == 0
    
    from openpyxl import load_workbook
    merged_wb = load_workbook(outcome.output_path, read_only=True, data_only=True)
    merged_ws = merged_wb.active
    rows = list(merged_ws.iter_rows(values_only=True))
    assert rows[0] == ("序号", "标题", "播放量")
    assert rows[1] == (1, "视频1", 100)
    assert rows[2] == (2, "视频2", 200)
    assert rows[3] == (3, "视频3", 300)

def test_merge_strict_mismatch(tmp_path):
    f1 = tmp_path / "f1_kw.xlsx"
    f2 = tmp_path / "f2_kw.xlsx"
    create_xlsx_file(f1, ["序号", "标题", "播放量"], [[1, "视频1", 100]])
    create_xlsx_file(f2, ["序号", "标题", "点赞量"], [[1, "视频2", 50]])
    
    outcome = merge_xlsx_files(tmp_path, keyword="kw", platform="test_run", schema_mode="strict")
    assert outcome.status == RunStatus.PARTIAL
    assert outcome.stats.success_count == 1
    assert outcome.stats.failed_count == 1
    assert outcome.stats.skipped_count == 0
    assert len(outcome.errors) == 1
    assert outcome.errors[0].code == "XLSX_SCHEMA_MISMATCH"

def test_merge_union_schema(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    f2 = tmp_path / "f2.xlsx"
    create_xlsx_file(f1, ["序号", "标题", "播放量"], [[1, "视频1", 100]])
    create_xlsx_file(f2, ["序号", "标题", "点赞量"], [[1, "视频2", 50]])
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run", schema_mode="union_schema")
    assert outcome.status == RunStatus.SUCCEEDED
    assert outcome.stats.success_count == 2
    assert outcome.stats.failed_count == 0
    
    from openpyxl import load_workbook
    merged_wb = load_workbook(outcome.output_path, read_only=True, data_only=True)
    merged_ws = merged_wb.active
    rows = list(merged_ws.iter_rows(values_only=True))
    assert rows[0] == ("序号", "标题", "播放量", "点赞量")
    assert rows[1] == (1, "视频1", 100, None)
    assert rows[2] == (2, "视频2", None, 50)

def test_merge_duplicate_headers(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    create_xlsx_file(f1, ["序号", "标题", "标题"], [[1, "A", "B"]])
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run")
    assert outcome.status == RunStatus.FAILED
    assert any(e.code == "XLSX_DUPLICATE_HEADER" for e in outcome.errors)

def test_merge_empty_worksheet(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    create_xlsx_file(f1, [], [])
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run")
    assert outcome.status == RunStatus.FAILED
    assert any(e.code == "XLSX_WORKSHEET_EMPTY" for e in outcome.errors)

def test_merge_corrupted_file(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    with open(f1, "w") as f:
        f.write("corrupted data")
        
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run")
    assert outcome.status == RunStatus.FAILED
    assert any(e.code == "XLSX_FILE_UNREADABLE" for e in outcome.errors)

def test_merge_formulas(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "数A", "数B", "和"])
    ws.append([1, 10, 20, "=SUM(B2:C2)"])
    wb.save(f1)
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run")
    assert outcome.status == RunStatus.SUCCEEDED

def test_merge_cancelled(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    create_xlsx_file(f1, ["序号", "标题"], [[1, "视频1"]])
    
    stop_event = threading.Event()
    stop_event.set()
    
    outcome = merge_xlsx_files(tmp_path, keyword="", platform="test_run", stop_event=stop_event)
    assert outcome.status == RunStatus.CANCELLED
    assert len(outcome.errors) == 1
    assert outcome.errors[0].code == "RUN_CANCELLED"

def test_merge_concurrency_isolation(tmp_path):
    f1 = tmp_path / "f1.xlsx"
    create_xlsx_file(f1, ["序号", "标题"], [[1, "A"]])
    
    outcomes = []
    def run_merge(rid):
        outcomes.append(merge_xlsx_files(tmp_path, keyword="", platform="test_run", run_id=rid))
        
    t1 = threading.Thread(target=run_merge, args=("run_A",))
    t2 = threading.Thread(target=run_merge, args=("run_B",))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    
    assert len(outcomes) == 2
    assert outcomes[0].status == RunStatus.SUCCEEDED
    assert outcomes[1].status == RunStatus.SUCCEEDED
    assert "run_A" in outcomes[0].output_path or "run_A" in outcomes[1].output_path
    assert "run_B" in outcomes[0].output_path or "run_B" in outcomes[1].output_path
