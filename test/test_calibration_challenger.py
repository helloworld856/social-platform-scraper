from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.tools.calibration import STATUS_EMPTY_RESULT, load_config, main, parse_games_definition, parse_platforms, run_calibration_task


def create_mock_excel(file_path: Path, values: list[str]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "视频信息"
    sheet.append(["搜索词", "序号", "视频标题", "播放量", "点赞数", "发布时间", "视频链接"])
    for index, value in enumerate(values, 1):
        sheet.append(["kw", index, "title", "1", "1", "2026-06-18", value])
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)
    workbook.close()


def make_track(
    *,
    platform: str = "youtube",
    language: str = "en",
    official_keywords: list[str] | None = None,
    candidate_keywords: list[str] | None = None,
) -> dict[str, object]:
    return {
        "platform": platform,
        "language": language,
        "official_keywords": official_keywords or ["base"],
        "candidate_keywords": candidate_keywords or [],
    }


def make_game(*tracks: dict[str, object]) -> dict[str, object]:
    return {
        "name": "Game A",
        "tracks": list(tracks),
    }


def test_parse_platforms_keeps_unknown_for_cli_reporting():
    assert parse_platforms("youtube, unknown_platform") == ["youtube", "unknown_platform"]


def test_sample_calibration_config_uses_new_track_schema():
    config = load_config("config/calibration_config.json")

    assert config["platforms"] == ["youtube", "tiktok", "x_twitter"]
    assert config["x_twitter"]["x_search_tab"] == "latest"
    assert config["games"][0]["tracks"][0]["platform"] == "youtube"
    assert config["games"][0]["tracks"][0]["language"] == "en"
    assert config["games"][0]["tracks"][0]["official_keywords"] == ["Genshin Impact"]


def test_legacy_output_file_path_creates_output_calibration_run_dir(tmp_path):
    official_excel = tmp_path / "official.xlsx"
    create_mock_excel(official_excel, [])

    def mock_youtube(*args, **kwargs):
        kwargs["finish_callback"](str(official_excel))

    config = {
        "platforms": ["youtube"],
        "time_period": {"days": 7},
        "youtube": {"api_keys": ["key"], "max_results": 10},
        "games": [make_game(make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=["kw1"]))],
    }

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_youtube):
        run_dir = Path(run_calibration_task(config, str(tmp_path / "report.md")))

    assert run_dir.parent.name == "calibration"
    assert (run_dir / "reports" / "keyword_collection_standardized.xlsx").exists()


def test_empty_official_result_remains_placeholder_row(tmp_path):
    official_excel = tmp_path / "official.xlsx"
    candidate_excel = tmp_path / "candidate.xlsx"
    create_mock_excel(official_excel, [])
    create_mock_excel(candidate_excel, ["https://www.youtube.com/watch?v=abcdefghijk"])

    def mock_youtube(*args, **kwargs):
        keyword = kwargs["keywords_list"][0]
        kwargs["finish_callback"](str(official_excel if keyword == "base" else candidate_excel))

    config = {
        "platforms": ["youtube"],
        "time_period": {"days": 7},
        "youtube": {"api_keys": ["key"], "max_results": 10},
        "games": [make_game(make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=["kw1"]))],
    }

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_youtube):
        run_dir = Path(run_calibration_task(config, str(tmp_path / "output")))

    raw_official = json.loads(next((run_dir / "raw").rglob("official_01.json")).read_text(encoding="utf-8"))
    assert raw_official["status"] == STATUS_EMPTY_RESULT

    workbook = openpyxl.load_workbook(run_dir / "reports" / "keyword_collection_standardized.xlsx")
    sheet = workbook["标准化采集数据"]
    assert sheet.cell(row=2, column=4).value == "official"
    assert sheet.cell(row=2, column=12).value == STATUS_EMPTY_RESULT
    workbook.close()


def test_main_handles_string_days_and_invalid_days(tmp_path):
    valid_config = tmp_path / "valid.json"
    invalid_config = tmp_path / "invalid.json"
    valid_config.write_text(
        json.dumps(
            {
                "games": [{"name": "Game A", "tracks": [make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=[])]}],
                "time_period": {"days": "7"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    invalid_config.write_text(
        json.dumps(
            {
                "games": [{"name": "Game A", "tracks": [make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=[])]}],
                "time_period": {"days": "seven"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    with (
        patch.object(sys, "argv", ["calibration.py", "--config", str(valid_config)]),
        patch("src.tools.calibration.run_calibration_task") as mock_run,
    ):
        main()
        mock_run.assert_called_once()

    with patch.object(sys, "argv", ["calibration.py", "--config", str(invalid_config)]):
        with pytest.raises(SystemExit) as excinfo:
            main()
        assert excinfo.value.code == 1


def test_window_validation_skips_youtube_key_when_youtube_track_is_not_active():
    pytest.importorskip("PyQt5")
    from PyQt5.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    from src.tools.windows import CalibrationToolWindow

    window = CalibrationToolWindow()
    values = {
        "days": 7,
        "platforms": "tiktok, x_twitter",
        "youtube_api_keys": "",
        "youtube_max_results": 10,
        "tiktok_max_videos": 10,
        "x_max_scrolls": 2,
        "x_search_tab": "latest",
        "cdp_url": "http://localhost:9222",
        "output_path": "output/calibration",
        "games_definition": json.dumps(
            [
                make_game(
                    make_track(platform="tiktok", language="ja", official_keywords=["原神"], candidate_keywords=["原神 攻略"]),
                    make_track(platform="x_twitter", language="en", official_keywords=["Genshin Impact"], candidate_keywords=["genshin build"]),
                )
            ],
            ensure_ascii=False,
        ),
    }
    window.validate_values(values)
    assert app is not None


def test_window_validation_rejects_when_selected_platforms_match_no_tracks():
    pytest.importorskip("PyQt5")
    from PyQt5.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    from src.tools.windows import CalibrationToolWindow

    window = CalibrationToolWindow()
    values = {
        "days": 7,
        "platforms": "x_twitter",
        "youtube_api_keys": "",
        "youtube_max_results": 10,
        "tiktok_max_videos": 10,
        "x_max_scrolls": 2,
        "x_search_tab": "latest",
        "cdp_url": "http://localhost:9222",
        "output_path": "output/calibration",
        "games_definition": json.dumps(
            [make_game(make_track(platform="youtube", language="en", official_keywords=["Genshin Impact"], candidate_keywords=["genshin build"]))],
            ensure_ascii=False,
        ),
    }

    with pytest.raises(ValueError, match="track 不匹配"):
        window.validate_values(values)
    assert app is not None


def test_games_editor_widget_emits_new_track_definition_json():
    pytest.importorskip("PyQt5")
    from PyQt5.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    from src.tools.windows import CalibrationToolWindow

    window = CalibrationToolWindow()
    editor = window.widgets["games_definition"]
    serialized = editor.text()
    parsed = parse_games_definition(serialized)

    assert parsed[0]["name"] == "Genshin Impact"
    assert parsed[0]["tracks"][0]["platform"] == "youtube"
    assert parsed[0]["tracks"][0]["language"] == "en"
    assert parsed[0]["tracks"][0]["official_keywords"] == ["Genshin Impact"]
    assert "Genshin guide" in parsed[0]["tracks"][0]["candidate_keywords"]
    assert app is not None


def test_calibration_window_persists_main_form_values(tmp_path, monkeypatch):
    pytest.importorskip("PyQt5")
    from PyQt5.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    import src.core.config_store as config_store
    from src.tools.windows import CalibrationToolWindow

    monkeypatch.setattr(config_store, "get_config_dir", lambda: tmp_path)

    first_window = CalibrationToolWindow()
    first_window.widgets["days"].setValue(30)
    first_window.widgets["platforms"].setText("youtube, x_twitter")
    first_window.widgets["youtube_api_keys"].setPlainText("key-a\nkey-b")
    first_window.widgets["output_path"].setText("output/custom_calibration")
    first_window.widgets["games_definition"].setText(
        json.dumps(
            [make_game(make_track(platform="youtube", language="en", official_keywords=["Zenless Zone Zero"], candidate_keywords=["zzz build"]))],
            ensure_ascii=False,
        )
    )
    values = first_window.collect_values()
    assert values is not None
    first_window._save_form_values(values)

    second_window = CalibrationToolWindow()
    assert second_window.widgets["days"].value() == 30
    assert second_window.widgets["platforms"].text() == "youtube, x_twitter"
    assert second_window.widgets["youtube_api_keys"].toPlainText() == "key-a\nkey-b"
    assert second_window.widgets["output_path"].text() == "output/custom_calibration"

    parsed = parse_games_definition(second_window.widgets["games_definition"].text())
    assert parsed[0]["tracks"][0]["official_keywords"] == ["Zenless Zone Zero"]
    assert app is not None
