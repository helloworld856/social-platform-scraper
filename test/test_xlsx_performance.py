import time
from openpyxl import load_workbook

from src.core.xlsx import MultiSheetXlsxWriter

def test_xlsx_performance_comparison(tmp_path):
    # Prepare test data
    fieldnames = ["col1", "col2", "col3", "col4", "col5"]
    rows = [{"col1": f"data_{i}_1", "col2": f"data_{i}_2", "col3": f"data_{i}_3", "col4": f"data_{i}_4", "col5": f"data_{i}_5"} for i in range(1000)]
    
    # 1. Baseline: MultiSheetXlsxWriter with autosave_every=1, writing line-by-line
    baseline_path = tmp_path / "baseline.xlsx"
    start_time = time.perf_counter()
    writer_base = MultiSheetXlsxWriter(str(baseline_path), {"Sheet1": fieldnames}, autosave_every=1)
    for row in rows:
        writer_base.writerow("Sheet1", row)
    baseline_duration = time.perf_counter() - start_time
    
    # Verify baseline file
    wb_base = load_workbook(baseline_path)
    ws_base = wb_base["Sheet1"]
    assert ws_base.max_row == 1001  # 1 header + 1000 rows
    assert ws_base.cell(row=2, column=1).value == "data_0_1"
    
    # 2. Optimized: MultiSheetXlsxWriter with autosave_every=500, writing in batches using writerows
    opt_path = tmp_path / "optimized.xlsx"
    start_time = time.perf_counter()
    writer_opt = MultiSheetXlsxWriter(str(opt_path), {"Sheet1": fieldnames}, autosave_every=500)
    writer_opt.writerows("Sheet1", rows)
    opt_duration = time.perf_counter() - start_time
    
    # Verify optimized file
    wb_opt = load_workbook(opt_path)
    ws_opt = wb_opt["Sheet1"]
    assert ws_opt.max_row == 1001
    assert ws_opt.cell(row=1001, column=5).value == "data_999_5"
    
    print(f"\nBaseline duration: {baseline_duration:.4f}s")
    print(f"Optimized duration: {opt_duration:.4f}s")
    
    speedup = baseline_duration / max(1e-9, opt_duration)
    print(f"Speedup: {speedup:.2f}x")
    
    # Assert at least 5x speedup
    assert speedup >= 5.0, f"Optimized version is only {speedup:.2f}x faster, expected >= 5.0x"
