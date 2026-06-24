from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.tools.calibration import (
    KEYWORD_ROLE_CANDIDATE,
    KEYWORD_ROLE_OFFICIAL,
    STANDARDIZED_HEADERS,
    STANDARDIZED_SHEET_NAME,
    STATUS_EMPTY_RESULT,
    STATUS_SUCCESS,
    STATUS_UNKNOWN_PLATFORM,
    extract_id_from_link,
    format_keyword_groups_text,
    parse_games_definition,
    parse_keyword_groups_text,
    parse_platforms,
    run_calibration_task,
    run_platform_spider,
    select_x_search_tab,
)


def create_mock_excel(file_path: Path, platform: str, rows: list[dict[str, str]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if platform == "x_twitter":
        sheet.title = "数据"
        sheet.append(["原始搜索词", "完整搜索语法", "序号", "推文内容", "浏览量", "点赞量", "转发量", "评论数", "发帖时间", "推文链接", "标签"])
        for index, row in enumerate(rows, 1):
            sheet.append(
                [
                    "kw",
                    "kw",
                    index,
                    row.get("title", row.get("description", "content")),
                    "10",
                    "1",
                    "1",
                    "1",
                    row.get("published_at", "2026-06-18"),
                    row.get("link", ""),
                    "tag",
                ]
            )
    else:
        sheet.title = "视频信息"
        sheet.append(["搜索词", "序号", "视频标题", "播放量", "点赞数", "发布时间", "视频链接"])
        for index, row in enumerate(rows, 1):
            sheet.append(
                [
                    "kw",
                    index,
                    row.get("title", "title"),
                    "10",
                    "1",
                    row.get("published_at", "2026-06-18"),
                    row.get("link", ""),
                ]
            )
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)
    workbook.close()


def read_standardized_rows(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[STANDARDIZED_SHEET_NAME]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
    rows: list[dict[str, str]] = []
    for row_number in range(2, sheet.max_row + 1):
        rows.append(
            {
                str(headers[index - 1]): "" if sheet.cell(row=row_number, column=index).value is None else str(sheet.cell(row=row_number, column=index).value)
                for index in range(1, sheet.max_column + 1)
            }
        )
    workbook.close()
    return rows


def make_track(
    *,
    platform: str = "youtube",
    language: str = "en",
    official_keywords: list[str] | None = None,
    candidate_keywords: list[str] | None = None,
) -> dict[str, object]:
    return {
        "platform": platform,
        "language": language,
        "official_keywords": official_keywords or ["base"],
        "candidate_keywords": candidate_keywords or [],
    }


def make_game(*tracks: dict[str, object]) -> dict[str, object]:
    return {
        "name": "Game A",
        "tracks": list(tracks),
    }


def test_parse_platforms_deduplicates_and_defaults():
    assert parse_platforms("youtube, tiktok, youtube, x_twitter") == ["youtube", "tiktok", "x_twitter"]
    assert parse_platforms("") == ["youtube", "tiktok", "x_twitter"]


def test_keyword_groups_text_helpers_round_trip():
    raw_text = "kw1, kw2\n\n词组A，词组B\nsolo"
    groups = parse_keyword_groups_text(raw_text)
    assert groups == [["kw1", "kw2"], ["词组A", "词组B"], ["solo"]]
    assert format_keyword_groups_text(groups) == "kw1, kw2\n词组A, 词组B\nsolo"


def test_parse_games_definition_supports_legacy_and_new_track_formats():
    legacy_definition = """
    # 注释会被忽略
    Game A | base
    kw1, kw2
    kw3
    """.strip()

    parsed_legacy = parse_games_definition(legacy_definition)
    assert parsed_legacy == [
        {
            "name": "Game A",
            "tracks": [
                make_track(platform="youtube", language="default", official_keywords=["base"], candidate_keywords=["kw1", "kw2", "kw3"]),
                make_track(platform="tiktok", language="default", official_keywords=["base"], candidate_keywords=["kw1", "kw2", "kw3"]),
                make_track(platform="x_twitter", language="default", official_keywords=["base"], candidate_keywords=["kw1", "kw2", "kw3"]),
            ],
        }
    ]

    json_definition = json.dumps(
        [
            make_game(
                make_track(platform="youtube", language="en", official_keywords=["base-c"], candidate_keywords=["alpha", "beta"]),
                make_track(platform="tiktok", language="ja", official_keywords=["基准词"], candidate_keywords=["词组A", "词组B"]),
            )
        ],
        ensure_ascii=False,
    )
    assert parse_games_definition(json_definition) == [
        {
            "name": "Game A",
            "tracks": [
                make_track(platform="youtube", language="en", official_keywords=["base-c"], candidate_keywords=["alpha", "beta"]),
                make_track(platform="tiktok", language="ja", official_keywords=["基准词"], candidate_keywords=["词组A", "词组B"]),
            ],
        }
    ]


def test_parse_games_definition_rejects_invalid_block_header():
    with pytest.raises(ValueError, match="首行必须写成"):
        parse_games_definition("Game A\nkw1, kw2")


def test_extract_id_from_link_normalizes_platform_urls():
    assert extract_id_from_link("https://www.youtube.com/watch?v=abcdefghijk&feature=share", "youtube") == "abcdefghijk"
    assert extract_id_from_link("https://m.tiktok.com/v/7359934348222222222.html?foo=bar", "tiktok") == "7359934348222222222"
    assert extract_id_from_link("https://twitter.com/user/status/1234567890?s=20", "x_twitter") == "1234567890"


def test_run_platform_spider_reports_success_empty_and_unknown(tmp_path):
    success_excel = tmp_path / "success.xlsx"
    empty_excel = tmp_path / "empty.xlsx"
    create_mock_excel(success_excel, "youtube", [{"link": "https://www.youtube.com/watch?v=abcdefghijk", "title": "video a"}])
    create_mock_excel(empty_excel, "youtube", [])

    def mock_success(*args, **kwargs):
        kwargs["stats_callback"]({"scanned_count": 5, "written_count": 1, "hit_limit": True})
        kwargs["finish_callback"](str(success_excel))

    def mock_empty(*args, **kwargs):
        kwargs["stats_callback"]({"scanned_count": 3, "written_count": 0, "hit_limit": False})
        kwargs["finish_callback"](str(empty_excel))

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_success):
        result = run_platform_spider("youtube", "kw", "2026-06-01", "2026-06-08", {}, 7)
        assert result.status == STATUS_SUCCESS
        assert result.ids == {"abcdefghijk"}
        assert result.records[0]["title"] == "video a"
        assert result.output_path == str(success_excel)
        assert result.scanned_count == 5
        assert result.written_count == 1
        assert result.hit_limit is True

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_empty):
        result = run_platform_spider("youtube", "kw", "2026-06-01", "2026-06-08", {}, 7)
        assert result.status == STATUS_EMPTY_RESULT
        assert result.records == []
        assert result.scanned_count == 3
        assert result.written_count == 0
        assert result.hit_limit is False

    result = run_platform_spider("unknown_platform", "kw", "2026-06-01", "2026-06-08", {}, 7)
    assert result.status == STATUS_UNKNOWN_PLATFORM


def test_x_search_tab_helpers_support_latest_and_top():
    assert select_x_search_tab({"x_search_tab": "latest"}) == "latest"
    assert select_x_search_tab({"x_search_tab": "top"}) == "top"
    assert select_x_search_tab({"x_search_tab": "invalid"}) == "latest"

    source = Path("src/platforms/x_twitter/keyword.py").read_text(encoding="utf-8")
    assert '"latest": "live"' in source
    assert '"top": "top"' in source


def test_run_calibration_task_generates_standardized_bundle(tmp_path):
    official_excel = tmp_path / "official.xlsx"
    candidate_a_excel = tmp_path / "candidate_a.xlsx"
    candidate_b_excel = tmp_path / "candidate_b.xlsx"
    create_mock_excel(
        official_excel,
        "youtube",
        [
            {"link": "https://www.youtube.com/watch?v=aaaaaaaaaaa", "title": "official a"},
            {"link": "https://youtu.be/bbbbbbbbbbb", "title": "official b"},
        ],
    )
    create_mock_excel(
        candidate_a_excel,
        "youtube",
        [
            {"link": "https://www.youtube.com/watch?v=aaaaaaaaaaa", "title": "candidate a"},
            {"link": "https://www.youtube.com/watch?v=ccccccccccc", "title": "candidate c"},
        ],
    )
    create_mock_excel(
        candidate_b_excel,
        "youtube",
        [{"link": "https://www.youtube.com/watch?v=ddddddddddd", "title": "candidate d"}],
    )

    def mock_youtube(*args, **kwargs):
        keyword = kwargs["keywords_list"][0]
        stats_map = {
            "base": {"scanned_count": 2, "written_count": 2, "hit_limit": False},
            "kw1": {"scanned_count": 2, "written_count": 2, "hit_limit": False},
            "kw2": {"scanned_count": 3, "written_count": 1, "hit_limit": True},
        }
        path_map = {
            "base": official_excel,
            "kw1": candidate_a_excel,
            "kw2": candidate_b_excel,
        }
        kwargs["stats_callback"](stats_map[keyword])
        kwargs["finish_callback"](str(path_map[keyword]))

    config = {
        "platforms": ["youtube"],
        "time_period": {"days": 7},
        "youtube": {"api_keys": ["key"], "max_results": 10},
        "games": [make_game(make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=["kw1", "kw2"]))],
    }

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_youtube):
        run_dir = Path(run_calibration_task(config, str(tmp_path / "legacy_report.md")))

    assert (run_dir / "config_snapshot.json").exists()
    assert (run_dir / "environment_snapshot.json").exists()
    standardized_path = run_dir / "reports" / "keyword_collection_standardized.xlsx"
    assert standardized_path.exists()
    assert not (run_dir / "reports" / "calibration_report.md").exists()
    assert not (run_dir / "reports" / "calibration_report.csv").exists()

    official_paths = list((run_dir / "raw").rglob("official_01.json"))
    candidate_paths = sorted((run_dir / "raw").rglob("candidate_*.json"))
    assert len(official_paths) == 1
    assert len(candidate_paths) == 2

    workbook = openpyxl.load_workbook(standardized_path)
    assert workbook.sheetnames == [STANDARDIZED_SHEET_NAME]
    headers = [workbook[STANDARDIZED_SHEET_NAME].cell(row=1, column=index).value for index in range(1, len(STANDARDIZED_HEADERS) + 1)]
    workbook.close()
    assert headers == STANDARDIZED_HEADERS

    rows = read_standardized_rows(standardized_path)
    assert len(rows) == 5
    assert rows[0]["关键词角色"] == KEYWORD_ROLE_OFFICIAL
    assert rows[0]["关键词文本"] == "base"
    assert rows[0]["采集状态"] == STATUS_SUCCESS
    assert rows[2]["关键词角色"] == KEYWORD_ROLE_CANDIDATE
    assert rows[2]["关键词文本"] == "kw1"
    assert rows[2]["内容ID"] == "aaaaaaaaaaa"
    assert rows[4]["关键词文本"] == "kw2"
    assert rows[4]["内容ID"] == "ddddddddddd"

    snapshot = json.loads(candidate_paths[1].read_text(encoding="utf-8"))
    assert snapshot["keyword_role"] == KEYWORD_ROLE_CANDIDATE
    assert snapshot["keyword"] == "kw2"
    assert snapshot["record_count"] == 1
    assert snapshot["hit_limit"] is True

    env_snapshot = json.loads((run_dir / "environment_snapshot.json").read_text(encoding="utf-8"))
    assert env_snapshot["x_search_tab"] == "latest"
    assert env_snapshot["standardized_output_path"] == str(standardized_path)


def test_run_calibration_task_keeps_placeholder_rows_for_empty_results(tmp_path):
    official_excel = tmp_path / "official_empty.xlsx"
    candidate_excel = tmp_path / "candidate.xlsx"
    create_mock_excel(official_excel, "youtube", [])
    create_mock_excel(candidate_excel, "youtube", [{"link": "https://www.youtube.com/watch?v=abcdefghijk", "title": "candidate"}])

    def mock_youtube(*args, **kwargs):
        keyword = kwargs["keywords_list"][0]
        kwargs["finish_callback"](str(official_excel if keyword == "base" else candidate_excel))

    config = {
        "platforms": ["youtube"],
        "time_period": {"days": 7},
        "youtube": {"api_keys": ["key"], "max_results": 10},
        "games": [make_game(make_track(platform="youtube", language="en", official_keywords=["base"], candidate_keywords=["kw1"]))],
    }

    with patch("src.tools.calibration.run_youtube_spider", side_effect=mock_youtube):
        run_dir = Path(run_calibration_task(config, str(tmp_path / "output")))

    rows = read_standardized_rows(run_dir / "reports" / "keyword_collection_standardized.xlsx")
    assert len(rows) == 2
    assert rows[0]["关键词角色"] == KEYWORD_ROLE_OFFICIAL
    assert rows[0]["采集状态"] == STATUS_EMPTY_RESULT
    assert rows[0]["内容ID"] == ""
    assert rows[1]["关键词角色"] == KEYWORD_ROLE_CANDIDATE
    assert rows[1]["内容ID"] == "abcdefghijk"
