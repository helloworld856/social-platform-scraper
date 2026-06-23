from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

pytest.importorskip("PyQt5")

from PyQt5.QtWidgets import QApplication

from src.processing.ai_semantic_rule_judge import (
    PASS_RESULT_EMPTY,
    REVIEW_STATUS_EMPTY,
    build_row_content,
    extract_sheet_headers,
    run_semantic_rule_judge,
    validate_headers,
)
from src.processing.windows import AISemanticRuleJudgeWindow

_app = None


def _ensure_app():
    global _app
    if _app is None:
        _app = QApplication.instance() or QApplication([])
    return _app


def _create_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(["标题", "简介", "其他"])
    data_sheet.append(["Alpha", "hero intro", "x"])
    data_sheet.append(["Beta", "side intro", "y"])
    data_sheet.append([None, None, "z"])
    other_sheet = workbook.create_sheet("Meta")
    other_sheet.append(["保留"])
    other_sheet.append(["value"])
    workbook.save(path)


def test_validate_headers_rejects_empty_and_duplicate():
    with pytest.raises(ValueError, match="空表头"):
        validate_headers(["标题", "", "简介"])
    with pytest.raises(ValueError, match="重复表头"):
        validate_headers(["标题", "简介", " 标题 "])


def test_build_row_content_uses_sheet_order_not_click_order():
    headers = ["简介", "标题", "其他"]
    row_values = ["intro", "title", "ignored"]
    content = build_row_content(headers, row_values, ["标题", "简介"])
    assert content == "简介：intro\n\n标题：title"


def test_extract_sheet_headers_reads_selected_sheet(tmp_path):
    file_path = tmp_path / "source.xlsx"
    _create_workbook(file_path)

    headers = extract_sheet_headers(file_path, "Data")

    assert headers == ["标题", "简介", "其他"]


def test_window_loads_sheets_and_refreshes_columns(tmp_path):
    _ensure_app()
    file_path = tmp_path / "source.xlsx"
    _create_workbook(file_path)

    window = AISemanticRuleJudgeWindow()
    window.widgets["input_xlsx"].path_edit.setText(str(file_path))
    window._reload_sheet_names(str(file_path))

    combo = window.widgets["sheet_name"]
    assert combo.count() == 2
    assert combo.currentText() == "Data"
    assert [window.column_list.item(i).text() for i in range(window.column_list.count())] == ["标题", "简介", "其他"]

    combo.setCurrentText("Meta")
    assert [window.column_list.item(i).text() for i in range(window.column_list.count())] == ["保留"]
    window.deleteLater()


def test_run_semantic_rule_judge_preserves_workbook_and_appends_results(tmp_path, monkeypatch):
    file_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_workbook(file_path)

    def fake_run_pass_batch(rows, rule_text, temperature, max_workers, batch_label):
        if batch_label.startswith("首轮"):
            return {
                row.row_number: ("符合" if row.row_number == 2 else "不符合", f"首轮-{row.row_number}", "")
                for row in rows
            }
        return {
            row.row_number: ("符合", f"复判-{row.row_number}", "")
            for row in rows
        }

    monkeypatch.setattr("src.processing.ai_semantic_rule_judge.run_pass_batch", fake_run_pass_batch)

    result = run_semantic_rule_judge(
        file_path,
        output_path,
        sheet_name="Data",
        target_columns=["简介", "标题"],
        rule_text="是否和主角相关",
        row_limit=10,
        max_workers=2,
        save_every_batches=1,
        temperature=0.1,
        sleep_seconds=0.0,
    )

    assert result == str(output_path)
    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["Data", "Meta"]

    sheet = workbook["Data"]
    appended_headers = [sheet.cell(row=1, column=index).value for index in range(4, 13)]
    assert appended_headers == [
        "AI判定词",
        "AI判定列",
        "首轮判定",
        "首轮理由",
        "复判判定",
        "复判理由",
        "最终判定",
        "复核状态",
        "异常信息",
    ]

    assert sheet.cell(row=2, column=5).value == "标题、简介"
    assert sheet.cell(row=2, column=6).value == "符合"
    assert sheet.cell(row=2, column=8).value == "符合"
    assert sheet.cell(row=2, column=10).value == "符合"
    assert sheet.cell(row=2, column=11).value == "一致"
    assert sheet.cell(row=3, column=10).value == "待人工复核"
    assert sheet.cell(row=3, column=11).value == "待人工复核"
    assert sheet.cell(row=3, column=12).value in (None, "")

    assert sheet.cell(row=4, column=6).value == PASS_RESULT_EMPTY
    assert sheet.cell(row=4, column=8).value == PASS_RESULT_EMPTY
    assert sheet.cell(row=4, column=10).value == PASS_RESULT_EMPTY
    assert sheet.cell(row=4, column=11).value == REVIEW_STATUS_EMPTY
    assert workbook["Meta"].cell(row=2, column=1).value == "value"


def test_run_semantic_rule_judge_marks_failed_batches_for_manual_review(tmp_path, monkeypatch):
    file_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output_failed.xlsx"
    _create_workbook(file_path)

    def fake_run_pass_batch(rows, rule_text, temperature, max_workers, batch_label):
        if batch_label.startswith("首轮"):
            return {row.row_number: ("符合", "首轮通过", "") for row in rows}
        return {row.row_number: ("失败", "", "网络异常") for row in rows}

    monkeypatch.setattr("src.processing.ai_semantic_rule_judge.run_pass_batch", fake_run_pass_batch)

    run_semantic_rule_judge(
        file_path,
        output_path,
        sheet_name="Data",
        target_columns=["标题"],
        rule_text="测试失败回退",
        row_limit=10,
        max_workers=1,
        save_every_batches=1,
        temperature=0.1,
        sleep_seconds=0.0,
    )

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook["Data"]
    assert sheet.cell(row=2, column=10).value == "待人工复核"
    assert sheet.cell(row=2, column=11).value == "待人工复核"
    assert "网络异常" in sheet.cell(row=2, column=12).value
